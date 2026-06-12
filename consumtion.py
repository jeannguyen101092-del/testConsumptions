import base64
import io
import json
import re
import requests
import streamlit as st
import pandas as pd
from pdf2image import convert_from_bytes, pdfinfo_from_bytes
from google import genai
from google.genai import types

# BẮT BUỘC: Câu lệnh cấu hình trang phải nằm đầu tiên trong file Streamlit
st.set_page_config(
    page_title="PPJ Techpack AI - Management System",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ĐỒ HỌA HIGH-CONTRAST INDUSTRIAL LIGHT THEME (XÓA BỎ BÓNG TỐI, CỐ ĐỊNH CHỮ RÕ NÉT)
st.markdown("""
    <style>
    /* Ép toàn bộ nền ứng dụng về màu xám trắng phòng thí nghiệm sạch sẽ */
    .stApp { background-color: #F8FAFC !important; }
    
    /* Thiết kế thanh điều hướng Sidebar màu trắng tinh, chữ xanh đen tương phản */
    [data-testid="stSidebar"] { 
        background-color: #FFFFFF !important; 
        border-right: 1px solid #CBD5E1 !important;
        min-width: 320px; 
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p { 
        color: #1E293B !important; font-weight: 600; font-size: 13.5px;
    }
    
    /* Khung thương hiệu PPJ Group hiệu ứng Gradient cao cấp */
    .sidebar-brand-container {
        background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
        padding: 22px; border-radius: 14px; text-align: center; margin-bottom: 30px;
        box-shadow: 0 4px 14px rgba(37, 99, 235, 0.2);
    }
    .sidebar-brand-title { font-size: 24px; font-weight: 800; color: #FFFFFF; margin: 0; letter-spacing: 1px; }
    .sidebar-brand-subtitle { font-size: 11px; color: #BFDBFE; margin-top: 5px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }
    
    /* Thiết kế tiêu đề phân hệ lớn dạng dải màu Gradient hoành tráng */
    .component-title-box {
        background: linear-gradient(90deg, #1E3A8A 0%, #2563EB 100%);
        color: #FFFFFF !important; font-size: 16px; font-weight: 700; padding: 14px 20px;
        border-radius: 10px; margin-bottom: 25px; letter-spacing: 0.5px; text-transform: uppercase;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.1);
    }
    
    /* Thiết kế Khung Container (Card hoành tráng, có đổ bóng tách biệt không gian) */
    .card-container {
        background-color: #FFFFFF !important; border: 1px solid #E2E8F0 !important; border-radius: 14px !important;
        padding: 24px; margin-bottom: 25px; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.03) !important;
    }
    
    /* Lưới thông tin Metadata bọc khung xám nhẹ */
    .metric-grid-box {
        display: flex; gap: 25px; background: #F8FAFC; padding: 14px 20px; border-radius: 10px; border: 1px solid #E2E8F0; margin-bottom: 20px;
    }
    .metric-label { font-size: 11px; font-weight: 700; color: #64748B; margin: 0; text-transform: uppercase; letter-spacing: 0.5px; }
    .metric-value { font-size: 14px; font-weight: 700; color: #1E3A8A; margin: 3px 0 0 0; }
    
    /* Bộ khung chứa bảng thông số kỹ thuật mượt mà */
    .data-table-container {
        max-height: 420px; overflow-y: auto; border: 1px solid #CBD5E1; border-radius: 10px; margin-top: 12px; background: white;
    }
    
    /* Định dạng bảng dữ liệu dệt may công nghiệp */
    .industrial-table { width: 100%; border-collapse: collapse; text-align: left; }
    .industrial-table th {
        background-color: #F1F5F9 !important; color: #1E3A8A !important; font-weight: 700 !important; padding: 12px 16px; font-size: 13px; position: sticky; top: 0; z-index: 5; border-bottom: 2px solid #CBD5E1 !important;
    }
    .industrial-table td { padding: 11px 16px; border-bottom: 1px solid #E2E8F0; color: #334155 !important; font-size: 13px; }
    .industrial-table tr:hover { background-color: #F8FAFC !important; }
    
    /* Khung thông báo trạng thái rỗng (Hệ thống IDLE) */
    .idle-alert-box {
        background-color: #FFFBEB; border-left: 5px solid #F59E0B; padding: 16px 20px; border-radius: 4px 12px 12px 4px; color: #B45309; font-size: 13.5px; font-weight: 600;
    }
    
    /* Ép toàn bộ màu chữ của bong bóng Chatbot về màu xám đậm trên nền trắng để nhìn rõ 100% */
    [data-testid="stChatMessage"] { background-color: #FFFFFF !important; border: 1px solid #CBD5E1 !important; border-radius: 12px !important; box-shadow: 0 2px 5px rgba(0,0,0,0.02) !important; }
    [data-testid="stChatMessage"] p { color: #0F172A !important; font-size: 14px !important; font-weight: 500 !important; line-height: 1.6 !important; }
    
    /* Đồng bộ màu chữ cho văn bản Streamlit tiêu chuẩn */
    [data-testid="stMarkdownContainer"] p, [data-testid="stMarkdownContainer"] h5 { color: #1E293B !important; }
    </style>
""", unsafe_allow_html=True)
# Cấu hình cổng kết nối Master DB của Tập đoàn PPJ Group
SB_URL = "https://ewqqodsfvlvnrzsylawy.supabase.co"
SB_KEY = st.secrets.get("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImV3cXFvZHNmdmx2bnJ6c3lsYXd5Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzUxMTkyOTAsImV4cCI6MjA5MDY5NTI5MH0.BWPxOsyswBT5CLrZgluRC1F2x5EpU06oexUFyakGhyc")

def get_secure_gemini_key():
    """Hàm bảo mật trích xuất Token API chìa khóa phân tích từ bộ Secrets"""
    if "GEMINI_API_KEY" in st.secrets:
        return st.secrets["GEMINI_API_KEY"].strip()
    return None
def is_valid_jpeg(data):
    """Hàm kiểm tra dữ liệu binary có phải là file ảnh JPEG chuẩn hay không"""
    if not data or len(data) < 4:
        return False
    # File JPEG chuẩn bắt đầu bằng FF D8 và kết thúc bằng FF D9
    return data.startswith(b'\xff\xd8')

def save_to_supabase_techpack_table(payload_data, raw_file_bytes=None, file_name=""):
    """
    Hàm xử lý đồng bộ dữ liệu nạp kho của Chức năng 1.
    🎯 SỬA LỖI GỐC MÀN HÌNH ĐEN: Thêm bộ kiểm định cấu trúc ảnh (is_valid_jpeg) trước khi nạp.
    Chỉ cho phép đẩy dữ liệu RAW bytes chuẩn lên Supabase Storage bằng phương thức PUT.
    """
    try:
        style_name_db = payload_data.get("style_number_parsed", "").strip()
        if not style_name_db or style_name_db == "UNKNOWN":
            file_style_match = re.search(r'([a-zA-Z0-9]+-[a-zA-Z0-9]+)', str(file_name))
            if file_style_match:
                style_name_db = file_style_match.group(1).strip()
            else:
                style_name_db = str(file_name).split('.')[0].strip()
                
        style_name_db = style_name_db.upper()
        sketch_b64 = payload_data.get("sketch_image", "")
        public_image_url = ""
        image_data = None

        # 1. Luồng trích xuất dữ liệu hình ảnh phẳng từ tệp PDF bản vẽ kỹ thuật
        if raw_file_bytes and file_name.lower().endswith('.pdf'):
            try:
                import pdfplumber
                from pdf2image import convert_from_bytes, pdfinfo_from_bytes
                
                info_pdf = pdfinfo_from_bytes(raw_file_bytes)
                total_p = int(info_pdf.get("Pages", 1))
                pdf_images = convert_from_bytes(raw_file_bytes, dpi=90, first_page=1, last_page=total_p)
                
                detected_idx = int(payload_data.get("sketch_page_index_detected", 0))
                best_idx = detected_idx
                
                with pdfplumber.open(io.BytesIO(raw_file_bytes)) as pdf:
                    if 0 <= detected_idx < len(pdf.pages):
                        page_text = pdf.pages[detected_idx].extract_text() or ""
                        tech_words = ["WAIST", "HIP", "INSEAM", "THIGH", "RISE", "SPEC", "TARGET", "TOLERANCE", "SIZE"]
                        word_count = sum(1 for w in tech_words if w in page_text.upper())
                        
                        if word_count >= 4 or len(page_text) > 400:
                            min_text_len = 99999
                            for i in range(min(4, len(pdf.pages))):
                                txt = pdf.pages[i].extract_text() or ""
                                c_count = sum(1 for w in txt.upper() if w in tech_words)
                                if c_count < 3 and len(txt) < min_text_len:
                                    min_text_len = len(txt)
                                    best_idx = i
                
                if 0 <= best_idx < len(pdf_images):
                    img_buf = io.BytesIO()
                    pdf_images[best_idx].convert("RGB").save(img_buf, format="JPEG", quality=85)
                    image_data = img_buf.getvalue()
            except Exception as img_err:
                print(f"[IMAGE EXTRACT ERROR]: Thất bại khi cắt ảnh từ PDF -> {str(img_err)}")
                image_data = None

        # Nếu trích xuất PDF không ra dữ liệu, thử giải mã chuỗi Base64 gửi kèm từ FE
        if not image_data and sketch_b64:
            try:
                import base64
                # Loại bỏ phần tiền tố data:image/...;base64, nếu có
                if "," in sketch_b64:
                    sketch_b64 = sketch_b64.split(",")[1]
                image_data = base64.b64decode(sketch_b64)
            except Exception as b64_err:
                print(f"[BASE64 DECODE ERROR]: Chuỗi Base64 ảnh bị lỗi -> {str(b64_err)}")
                image_data = None

        # 2. ĐỂY TẬP TIN HÌNH ẢNH SẢN PHẨM LÊN SUPABASE STORAGE KHO_ANH (ĐÃ KIỂM ĐỊNH FILE CHUẨN)
        if image_data:
            # 🛑 CHẶN ĐỨNG HÀNH VI ĐẨY FILE LỖI: Kiểm tra dữ liệu byte ảnh có đúng định dạng JPEG không
            if not is_valid_jpeg(image_data):
                print(f"[CRITICAL WARNING] Huỷ upload! Biến image_data cho mã {style_name_db} KHÔNG phải dữ liệu ảnh JPEG hợp lệ (kích thước: {len(image_data)} bytes). Vui lòng kiểm tra lại file PDF đầu vào.")
            else:
                try:
                    storage_headers = {
                        "apikey": SB_KEY, 
                        "Authorization": f"Bearer {SB_KEY}",
                        "Content-Type": "image/jpeg",
                        "x-upsert": "true"  # Cho phép ghi đè khi sửa lỗi
                    }
                    style_clean_filename = re.sub(r'[^a-zA-Z0-9_-]', '', style_name_db).upper()
                    storage_url = f"{SB_URL.rstrip('/')}/storage/v1/object/kho_anh/{style_clean_filename}.jpg"
                    
                    # Đẩy dữ liệu thô (raw bytes) lên API Supabase bằng phương thức PUT
                    upload_res = requests.put(
                        storage_url, 
                        headers=storage_headers, 
                        data=image_data, 
                        timeout=20
                    )
                    
                    if 200 <= upload_res.status_code <= 299:
                        print(f"[STORAGE SUCCESS] Upload ảnh gốc thành công cho mã: {style_clean_filename}")
                        public_image_url = f"{SB_URL.rstrip('/')}/storage/v1/object/public/kho_anh/{style_clean_filename}.jpg"
                    else:
                        print(f"[STORAGE ERROR] Supabase từ chối file. Mã lỗi {upload_res.status_code}: {upload_res.text}")
                except Exception as storage_err: 
                    print(f"[STORAGE EXCEPTION] Mất kết nối tới máy chủ Storage: {str(storage_err)}")
        else:
            print(f"[WARN] Không tìm thấy bất kỳ dữ liệu hình ảnh nào (image_data rỗng) cho mã {style_name_db}")

        # 3. LUỒNG KÍCH HOẠT MẮT THẦN AI VISION: TRÍCH XUẤT CHUỒI ĐẶC TRƯNG HÌNH HỌC
        measurements_raw = payload_data.get("measurements", {})
        visual_description_str = f"GARMENT TYPE: {payload_data.get('category', 'Garment Pants')}. Specs profile summary: " + ", ".join([f"{k}:{v}" for k, v in list(measurements_raw.items())[:6]])
        
        # Chỉ chạy AI Vision khi file ảnh được xác định là ảnh chuẩn
        if image_data and is_valid_jpeg(image_data):
            gemini_key = get_secure_gemini_key()
            if gemini_key:
                try:
                    from google import genai
                    from google.genai import types
                    
                    client_db = genai.Client(api_key=gemini_key)
                    vision_prompt = """
                    Analyze this technical garment flat sketch in detail.
                    List all unique geometric attributes, structural silhouette, waistband closure type, front/back pockets layout, panel shapes, and stitch lines.
                    Output a single dense string of these visual characteristics for apparel similarity vector matching.
                    Do not include greetings, just return the raw dense characteristic description string.
                    """
                    vision_res = client_db.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=[
                            types.Part.from_bytes(data=image_data, mime_type='image/jpeg'),
                            vision_prompt
                        ]
                    )
                    if vision_res and vision_res.text:
                        visual_description_str = vision_res.text.strip()
                except Exception as ai_vision_err:
                    print(f"[AI VISION RE-EXTRACT ERROR]: {str(ai_vision_err)}")

        # 4. Đẩy gói dữ liệu sạch đồng bộ lên bảng thong_so_techpack của Supabase
        headers = {
            "apikey": SB_KEY, 
            "Authorization": f"Bearer {SB_KEY}",
            "Content-Type": "application/json", 
            "Prefer": "resolution=merge-duplicates"
        }
        insert_url = f"{SB_URL.rstrip('/')}/rest/v1/thong_so_techpack"
        clean_dict = {str(k).strip(): str(v).strip() for k, v in dict(measurements_raw).items()}

        db_payload = {
            "StyleName": style_name_db,
            "Buyer": payload_data.get("buyer"),
            "Category": payload_data.get("category"),
            "BaseSize": payload_data.get("base_size_name"),
            "DetailedMeasurements": clean_dict,
            "SketchURL": public_image_url,
            "sketch_vector": visual_description_str
        }
        
        response = requests.post(insert_url, headers=headers, json=[db_payload], timeout=15)
        
        if 200 <= response.status_code <= 299:
            print("[DB SUCCESS] Đồng bộ dữ liệu bảng thong_so_techpack thành công!")
            return True
        else:
            print(f"[DB ERROR] Lỗi đồng bộ database {response.status_code}: {response.text}")
            return False
            
    except Exception as e:
        import streamlit as st
        st.sidebar.error(f"Lỗi xử lý hệ thống nạp kho: {str(e)}")
        print(f"[CRITICAL ERROR] Toàn hệ thống nạp kho thất bại: {str(e)}")
        return False







def get_historical_fabric_consumption_from_db(search_keyword=None):
    """
    Hàm tra cứu kho dữ liệu san_pham lịch sử nâng cao.
    ✨ ĐÃ SỬA LỖI TRỐNG BẢNG BOM: Áp dụng tìm kiếm mờ chuỗi lõi, không chia cắt chữ/số làm mất hậu tố wash dệt may.
    """
    try:
        headers = {
            "apikey": SB_KEY, 
            "Authorization": f"Bearer {SB_KEY}"
        }
        url = f"{SB_URL.rstrip('/')}/rest/v1/san_pham"
        
        query_params = {
            "select": "style_name,article_name,consumption_type,material_size,uom,consumption_value,notes",
            "limit": 1000
        }
        
        if search_keyword:
            kw_raw = str(search_keyword).strip().upper()
            # Làm sạch ký tự đặc biệt nhưng giữ nguyên vẹn chuỗi dài mã hàng để Supabase quét chính xác
            kw_clean = re.sub(r'[^A-Z0-9]', '', kw_raw)
            
            if len(kw_clean) >= 5:
                # Tìm kiếm mờ thông minh bao quát cả mã gốc lẫn các biến thể wash rách rập phân xưởng
                or_filter = f"(style_name.ilike.*{kw_clean}*,article_name.ilike.*{kw_clean}*)"
            else:
                or_filter = f"(style_name.ilike.*{kw_raw}*,article_name.ilike.*{kw_raw}*)"
                
            query_params["or"] = or_filter
        
        response = requests.get(url, headers=headers, params=query_params, timeout=15)
        return response.json() if response.status_code == 200 else []
    except Exception: 
        return []


def get_techpack_spec_from_db(style_name_keyword=None):
    """
    Hàm cho phép AI tự động tra cứu thông số từ bảng thong_so_techpack.
    ✨ ĐÃ CHUẨN HÓA: Đảm bảo đồng bộ chính xác tên các trường dữ liệu để trả về cho Đoạn 3 hiển thị.
    """
    try:
        headers = {
            "apikey": SB_KEY, 
            "Authorization": f"Bearer {SB_KEY}"
        }
        url = f"{SB_URL.rstrip('/')}/rest/v1/thong_so_techpack"
        
        query_params = {
            "select": "StyleName,Buyer,Category,BaseSize,DetailedMeasurements,SketchURL,sketch_vector",
            "limit": 500
        }
        
        if style_name_keyword and str(style_name_keyword).strip().upper() != "UNKNOWN":
            clean_kw = str(style_name_keyword).strip()
            query_params["StyleName"] = f"ilike.*{clean_kw}*"
            
        response = requests.get(url, headers=headers, params=query_params, timeout=15)
        return response.json() if response.status_code == 200 else []
    except Exception:
        return []

def process_single_pdf_batch(file_bytes, file_name):
    """
    Hàm bóc tách dữ liệu kỹ thuật từ một file PDF độc lập.
    ✨ ĐÃ NÂNG CẤP ĐỊNH VỊ PHOM DÁNG: Ép AI Vision chỉ bốc trang hiển thị chiếc quần hoàn chỉnh (Front and Back full garment views).
    STRICTLY FORBIDDEN: Cấm tuyệt đối lấy các trang rã rập thân quần đơn lẻ, cụm chi tiết hoặc rập tách rời.
    """
    import time
    try:
        gemini_key = get_secure_gemini_key()
        if not gemini_key:
            return {"success": False, "error": "API Key cho Gemini đang bị thiếu trong Secrets."}
            
        client = genai.Client(api_key=gemini_key)
        info = pdfinfo_from_bytes(file_bytes)
        total_p = int(info.get("Pages", 1))
        
        pdf_parts_payload = []
        chat_images = convert_from_bytes(file_bytes, dpi=90, first_page=1, last_page=total_p)
        for page_img in chat_images:
            img_buf = io.BytesIO()
            page_img.convert("RGB").save(img_buf, format="JPEG", quality=75)
            pdf_parts_payload.append(types.Part.from_bytes(data=img_buf.getvalue(), mime_type='image/jpeg'))
            
        industrial_extraction_prompt = (
            "You are an expert Garment Specification Auditor at PPJ Group. Analyze all attached sheets page by page. "
            "1. Identify the core 'Base Size' / 'Sample Size' (e.g., written as 8-, 32, or Size M). "
            "2. Identify the Buyer name and Category. "
            "3. Find the exact 'Style ID' / 'Style Number' (e.g. 5765). "
            "4. FOR FUNCTION 3 (FULL SIZE MATRIX): Scan and extract the entire grading matrix table columns for ALL available sizes. "
            "5. CRITICAL VISUAL FLAT SKETCH LOCATE RULE: Scan all pages visually. You MUST find the exact PAGE INDEX (0-based) "
            "that contains the FULL BODY APPAREL FLAT SKETCH showing the entire completed garment (the whole pant/skort with front view and back view side-by-side or on the same page). "
            "STRICT DISQUALIFICATION RULES: "
            "- DO NOT select pages showing isolated technical pattern panels (e.g., just a single front panel leg or a single back panel leg cut out). "
            "- DO NOT select pages showing inner construction details, pocket bags, zippers, or sketches of components. "
            "We only want the complete product design presentation sketch page. "
            "Return a completely valid raw JSON string matching this schema (no markdown blocks): "
            "{"
            "  \"style_number_parsed\": \"string\","
            "  \"buyer\": \"string\","
            "  \"category\": \"string\","
            "  \"base_size_name\": \"string\","
            "  \"sketch_page_index_detected\": 0,"
            "  \"measurements\": {\"POM Description\": \"Value\"},"
            "  \"full_size_matrix\": {\"POM Description\": {\"Size_Name\": \"Value\"}}"
            "}"
        )
        
        pdf_parts_payload.append(industrial_extraction_prompt)
        
        response = None
        for attempt in range(3):
            try:
                response = client.models.generate_content(
                    model='gemini-2.5-flash', 
                    contents=pdf_parts_payload,
                    config={"response_mime_type": "application/json"}
                )
                if response and response.text: break
            except Exception as ai_err:
                if "503" in str(ai_err) or "UNAVAILABLE" in str(ai_err):
                    time.sleep((attempt + 1) * 2)
                    continue
                else:
                    return {"success": False, "error": f"Lỗi cổng truyền: {str(ai_err)}"}
                    
        if not response or not response.text:
            return {"success": False, "error": "Mô hình không phản hồi văn bản."}
            
        clean_json = response.text.strip().replace("```json", "").replace("```", "").strip()
        parsed_data = json.loads(clean_json)
        
        extracted_sketch_bytes = None
        detected_idx = int(parsed_data.get("sketch_page_index_detected", 0))
        if 0 <= detected_idx < len(chat_images):
            b_buf = io.BytesIO()
            chat_images[detected_idx].convert("RGB").save(b_buf, format="JPEG", quality=90)
            extracted_sketch_bytes = b_buf.getvalue()
            
        success_db = save_to_supabase_techpack_table(parsed_data, raw_file_bytes=file_bytes, file_name=file_name)
        
        output_payload = {
            "style_number_parsed": parsed_data.get("style_number_parsed", "UNKNOWN"),
            "buyer": parsed_data.get("buyer", "UNKNOWN BUYER"),
            "category": parsed_data.get("category", "GARMENT"),
            "base_size_name": parsed_data.get("base_size_name", "32"),
            "measurements": parsed_data.get("measurements", {}),
            "full_size_matrix": parsed_data.get("full_size_matrix", {})
        }
        
        return {
            "success": True,
            "data": output_payload, 
            "style_id": output_payload["style_number_parsed"],
            "buyer": output_payload["buyer"],
            "category": output_payload["category"],
            "size": output_payload["base_size_name"],
            "measurements": output_payload["measurements"], 
            "sketch_bytes": extracted_sketch_bytes, 
            "error": None if success_db else "Lỗi ghi đồng bộ dữ liệu lên cơ sở dữ liệu"
        }
    except Exception as e:
        return {"success": False, "error": f"Lỗi bóc tách PDF: {str(e)}"}









# PHASE 5: USER INTERFACE STRUCTURE & AUTOMATION FACTORY 
# =============================================================================
with st.sidebar:
    st.markdown("""
        <div class="sidebar-brand-container">
            <div class="sidebar-brand-title">PPJ GROUP</div>
            <div class="sidebar-brand-subtitle">TECHPACK MANAGEMENT CORE AI</div>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("<p style='font-size:11px; font-weight:700; color:#64748B; margin: 15px 0 5px 5px; letter-spacing:0.5px;'>🏭 AUTOMATION FACTORY</p>", unsafe_allow_html=True)
    
    # ĐÃ ĐỒNG BỘ: Đảm bảo khớp hoàn toàn các nhãn chức năng
    menu_selection = st.radio(
        label="Chức năng hệ thống",
        options=["📊 Upload Techpack", "🔄 Pattern Spec Comparison", "🧵 BOM & Consumption Matrix","🛒 Purchase Consumption"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.success("DATABASE ACCESS: SECURED")
    st.info("ANALYTICS ENGINE: COMPLY")

if "processed_styles" not in st.session_state:
    st.session_state["processed_styles"] = {}


if menu_selection == "📊 Upload Techpack":
    import base64
    import concurrent.futures

    st.markdown('<div class="component-title-box">📊 MULTI-BATCH GARMENT SPECIFICATION MATRIX</div>', unsafe_allow_html=True)
    
    st.markdown("""<div class="card-container"><div class="card-section-header">📥 INGESTION ENGINE</div>
    <p style="color: #64748B; font-size:13px; margin:0 0 15px 0;">Hệ thống tự động cắt trang, khử nhiễu đồ họa phẳng và gọi API mạng nơ-ron tích hợp để bóc tách thông số hàng loạt.</p></div>""", unsafe_allow_html=True)
    
    uploaded_files = st.file_uploader("Upload Techpack PDFs Here", type=["pdf"], accept_multiple_files=True, key="bulk_techpack_pdf_uploader", label_visibility="collapsed")
    
    if uploaded_files:
        files_to_render = []
        files_need_processing = [f for f in uploaded_files if f.name not in st.session_state["processed_styles"]]
        
        if files_need_processing:
            if st.button(f"🚀 KÍCH HOẠT SỐ HÓA ĐA LUỒNG SONG SONG ({len(files_need_processing)} FILE MỚI)", use_container_width=True, type="primary"):
                status_text = st.empty()
                progress_bar = st.progress(0)
                total_new_files = len(files_need_processing)
                
                def thread_worker(file_obj):
                    try:
                        f_bytes = file_obj.getvalue()
                        res = process_single_pdf_batch(f_bytes, file_obj.name)
                        return {
                            "file_name": file_obj.name, 
                            "success": res.get("success", False), 
                            "style_id": res.get("style_id", "UNKNOWN"),
                            "buyer": res.get("buyer", "UNKNOWN BUYER"),
                            "category": res.get("category", "GARMENT"),
                            "size": res.get("size", "32"),
                            "measurements": res.get("measurements", {}),
                            "sketch_bytes": res.get("sketch_bytes", None),
                            "error": res.get("error", None),
                            "raw_bytes": f_bytes  
                        }
                    except Exception as e:
                        return {"file_name": file_obj.name, "success": False, "error": str(e), "raw_bytes": None}

                with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
                    future_to_file = {executor.submit(thread_worker, f): f.name for f in files_need_processing}
                    
                    for idx, future in enumerate(concurrent.futures.as_completed(future_to_file)):
                        f_name = future_to_file[future]
                        try:
                            task_res = future.result()
                            if task_res.get("success") == True:
                                # ĐỒNG BỘ TIỀN TỐ ẢNH BASE64: Ép trình duyệt tự động bung hình rập nét mảnh lớn lập tức
                                s_bytes = task_res.get("sketch_bytes")
                                img_base64_str = f"data:image/jpeg;base64,{base64.b64encode(s_bytes).decode('utf-8')}" if s_bytes else ""
                                
                                mock_data = {
                                    "style_number_parsed": task_res.get("style_id"),
                                    "buyer": task_res.get("buyer"), 
                                    "category": task_res.get("category"),
                                    "base_size_name": task_res.get("size"),
                                    "measurements": task_res.get("measurements", {}), 
                                    "sketch_image": img_base64_str, 
                                    "_raw_file_bytes": task_res["raw_bytes"] 
                                }
                                st.session_state["processed_styles"][f_name] = mock_data
                            else:
                                st.error(f"FAIL ENGINE [{f_name}]: {task_res.get('error')}")
                        except Exception as exc:
                            st.error(f"CRITICAL CRASH [{f_name}]: {str(exc)}")
                        
                        completed = idx + 1
                        progress_bar.progress(completed / total_new_files)
                        status_text.text(f"⚡ Core AI đang xử lý: {completed}/{total_new_files} tệp ({f_name})...")
                
                status_text.empty()
                progress_bar.empty()
                st.success("🎉 Số hóa dữ liệu thành công! Hãy kiểm tra bảng thông số bên dưới trước khi bấm lưu.")
        for file in uploaded_files:
            if file.name in st.session_state["processed_styles"]:
                files_to_render.append(file.name)

        if files_to_render:
            st.markdown("<br>", unsafe_allow_html=True)
            
            if st.button("💾 LƯU TOÀN BỘ DỮ LIỆU ĐÃ SỐ HÓA VÀO MASTER DB", key="bulk_save_all_btn", type="primary", use_container_width=True):
                success_count = 0
                with st.spinner("Đang đồng bộ cổng dữ liệu nhị phân hàng loạt lên Supabase Cloud..."):
                    for f_name in files_to_render:
                        style_data = st.session_state["processed_styles"][f_name]
                        raw_bytes_backup = style_data.get("_raw_file_bytes", None)
                        if save_to_supabase_techpack_table(payload_data=style_data, raw_file_bytes=raw_bytes_backup, file_name=f_name): 
                            success_count += 1
                st.success(f"🎉 PATTERN DATA PIPELINE: Đã bóc tách ảnh Sketch sạch và lưu trữ thành công {success_count}/{len(files_to_render)} mã hàng vào Database!")
            
            st.markdown("---")
            st.markdown("### 📋 KẾT QUẢ SỐ HÓA HÌNH HỌC VÀ THÔNG SỐ SẢN XUẤT")

            cols = st.columns(2)
            for idx, f_name in enumerate(files_to_render):
                col_target = cols[idx % 2]
                data = st.session_state["processed_styles"][f_name]
                with col_target:
                    st.markdown(f"""<div class="card-container"><div class="tech-card-header">{data.get('style_number_parsed')}</div>
                        <div class="metric-grid-box"><div><p class="metric-label">BUYER</p><p class="metric-value">{data.get('buyer')}</p></div>
                        <div><p class="metric-label">PRODUCT LINE</p><p class="metric-value">{data.get('category')}</p></div>
                        <div><p class="metric-label">BASE SIZE</p><p class="metric-value">{data.get('base_size_name')}</p></div></div></div>""", unsafe_allow_html=True)
                    
                    sub_col1, sub_col2 = st.columns([1.2, 0.8])
                    with sub_col1:
                        st.markdown("<p style='font-weight:700; font-size:12px; color:#1E293B;'>📋 SPECIFICATION DATA GRID</p>", unsafe_allow_html=True)
                        table_html = '<div class="data-table-container"><table class="industrial-table"><thead><tr><th>Point of Measurement</th><th>Target Spec</th></tr></thead><tbody>'
                        for k, v in data.get("measurements", {}).items():
                            table_html += f"<tr><td>{k}</td><td>{v}</td></tr>"
                        table_html += "</tbody></table></div>"
                        st.markdown(table_html, unsafe_allow_html=True)
                    with sub_col2:
                        st.markdown("<p style='font-weight:700; font-size:12px; color:#1E293B;'>📐 GARMENT FLAT SKETCH</p>", unsafe_allow_html=True)
                        # SỬA TRIỆT ĐỂ: Gọi biến trực tiếp chứa thẻ định vị, xóa hoàn toàn dấu ngoặc kép gây lỗi thô chữ
                        if data.get("sketch_image") and data["sketch_image"] != "":
                            try:
                                st.image(data["sketch_image"], use_container_width=True)
                            except Exception:
                                st.info("Hệ thống đang tải cổng ảnh vẽ phẳng kĩ thuật...")
                    st.markdown("<br><hr style='border-color:#E2E8F0;'><br>", unsafe_allow_html=True)
    else:
        st.markdown('<div class="idle-alert-box">⚠️ INITIALIZATION SYSTEM IDLE: Hiện tại chưa có tệp dữ liệu Techpack nào được nạp vào hệ thống để AI khởi chạy mô hình.</div>', unsafe_allow_html=True)







# -----------------------------------------------------------------------------
# CHỨC NĂNG 2: ĐỐI CHIẾU SO SÁNH HAI MÃ RẬP KHÁC NHAU (PATTERN SPEC COMPARISON)
# -----------------------------------------------------------------------------
elif menu_selection == "🔄 Pattern Spec Comparison":
    st.markdown('<div class="component-title-box">🔄 DIFFERENTIAL GEOMETRY & DELTA SPEC EVALUATOR</div>', unsafe_allow_html=True)
    st.markdown('<div class="card-container"><div class="card-section-header">🔍 CONFIGURATION SELECTION</div><p style="color: #64748B; font-size:13px; margin:0 0 15px 0;">Tải lên hai tệp bản vẽ kỹ thuật dệt may độc lập để tiến hành lập luận so sánh và tính toán toán học các khoảng chênh lệch rập mẫu.</p></div>', unsafe_allow_html=True)
    
    sc1, sc2 = st.columns(2)
    with sc1: file1 = st.file_uploader("Chọn file mẫu Techpack Gốc (File A)", type=["pdf"], key="f1")
    with sc2: file2 = st.file_uploader("Chọn file mẫu Techpack Sửa đổi (File B)", type=["pdf"], key="f2")
    
    if file1 and file2:
        # --- THUẬT TOÁN CÔ LẬP BỘ NHỚ ĐỆM TUYỆT ĐỐI CHỐNG LỆCH CỘT N/A ---
        if st.session_state.get("spec_last_f1") != file1.name or "spec_data_a" not in st.session_state:
            res1 = process_single_pdf_batch(file1.getvalue(), file1.name)
            if res1.get("success") and "data" in res1:
                st.session_state["spec_data_a"] = res1["data"]
                st.session_state["spec_last_f1"] = file1.name
            else:
                st.error(f"❌ Lỗi phân tích File A: {res1.get('error', 'Không có phản hồi')}")
                st.session_state.pop("spec_data_a", None)
                
        if st.session_state.get("spec_last_f2") != file2.name or "spec_data_b" not in st.session_state:
            res2 = process_single_pdf_batch(file2.getvalue(), file2.name)
            if res2.get("success") and "data" in res2:
                st.session_state["spec_data_b"] = res2["data"]
                st.session_state["spec_last_f2"] = file2.name
            else:
                st.error(f"❌ Lỗi phân tích File B: {res2.get('error', 'Không có phản hồi')}")
                st.session_state.pop("spec_data_b", None)
            
        d1 = st.session_state.get("spec_data_a")
        d2 = st.session_state.get("spec_data_b")
        
        if d1 and d2:
            style_a = d1.get('style_number_parsed', 'Mẫu A')
            style_b = d2.get('style_number_parsed', 'Mẫu B')
            
            # Gán nhãn phân biệt thông minh nếu người dùng upload cùng một mã thiết kế
            if style_a == style_b:
                lbl_a = f"Mẫu A ({style_a}-Gốc) [{d1.get('base_size_name','32').strip()}]"
                lbl_b = f"Mẫu B ({style_b}-Sửa) [{d2.get('base_size_name','32').strip()}]"
            else:
                lbl_a = f"Mẫu A ({style_a}) [{d1.get('base_size_name','32').strip()}]"
                lbl_b = f"Mẫu B ({style_b}) [{d2.get('base_size_name','32').strip()}]"
                
            st.info(f"⚙️ **ĐANG ĐỐI CHIẾU MA TRẬN PHÁT TRIỂN:** {lbl_a} ↔️ {lbl_b}")
            
            def clean_num(v):
                if not v or str(v).strip().upper() in ["N/A", ""]: return 0.0
                try:
                    s = str(v).replace("INCH", "").strip()
                    if " " in s:
                        p = s.split()
                        whole = float(p[0])
                        frac = p[1].split('/')
                        return whole + (float(frac[0]) / float(frac[1]))
                    return float(s.split('/')[0]) / float(s.split('/')[1]) if "/" in s else float(s)
                except:
                    import re
                    nums = re.findall(r"[-+]?\d*\.\d+|\d+", str(v))
                    return float(nums[0]) if nums else 0.0

            def extract_pom_code(pom_str):
                import re
                if not pom_str: return ""
                match = re.search(r'([A-Za-z]{2,4}-\d{3})', str(pom_str))
                return match.group(1).upper() if match else str(pom_str).lower().strip()

            df_a = pd.DataFrame(list(d1["measurements"].items()), columns=['raw_pom_a', lbl_a])
            df_b = pd.DataFrame(list(d2["measurements"].items()), columns=['raw_pom_b', lbl_b])
            
            df_a['pom_code'] = df_a['raw_pom_a'].apply(extract_pom_code)
            df_b['pom_code'] = df_b['raw_pom_b'].apply(extract_pom_code)
            
            df_a['seq'] = df_a.groupby('pom_code').cumcount()
            df_b['seq'] = df_b.groupby('pom_code').cumcount()
            
            df_res = pd.merge(df_a, df_b, on=['pom_code', 'seq'], how='outer').fillna("N/A").sort_values(['pom_code', 'seq'])
            table_body_html = ""
            compare_rows_for_df = []
            
            for _, r in df_res.iterrows():
                display_pom = r['raw_pom_a'] if r['raw_pom_a'] != "N/A" else r['raw_pom_b']
                val1, val2 = r[lbl_a], r[lbl_b]
                
                delta = round(clean_num(val2) - clean_num(val1), 3) if val1 != "N/A" and val2 != "N/A" else "N/A"
                compare_rows_for_df.append({"Vị trí đo (POM)": display_pom, lbl_a: val1, lbl_b: val2, "Sai lệch (Delta)": delta})
                
                if delta == "N/A":
                    style, txt = "color:#94A3B8; font-style:italic;", "N/A"
                elif delta > 0:
                    style, txt = "background:rgba(16,185,129,0.15); color:#166534; font-weight:700; padding:2px 8px; border-radius:4px; font-size:12px; border:1px solid #BBF7D0;", f"+{delta}"
                elif delta < 0:
                    style, txt = "background:rgba(239,68,68,0.15); color:#991B1B; font-weight:700; padding:2px 8px; border-radius:4px; font-size:12px; border:1px solid #FECACA;", f"{delta}"
                else:
                    style, txt = "color:#64748B; font-size:12px;", "0.00"
                
                table_body_html += f"<tr style='background:#FFF;'><td style='padding:10px 14px; border-bottom:1px solid #E2E8F0; font-weight:600; color:#1E293B;'>{display_pom}</td><td style='padding:10px 14px; border-bottom:1px solid #E2E8F0; color:#334155;'>{val1}</td><td style='padding:10px 14px; border-bottom:1px solid #E2E8F0; color:#334155;'>{val2}</td><td style='padding:10px 14px; border-bottom:1px solid #E2E8F0; text-align:center;'><span style='{style}'>{txt}</span></td></tr>"
            
            full_table_render = f"""
            <div style="max-height: 460px; overflow-y: auto; border: 1px solid #CBD5E1; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.02); margin-top: 15px;">
                <table style="width: 100%; border-collapse: collapse; text-align: left; font-family: sans-serif;">
                    <thead>
                        <tr style="background: linear-gradient(90deg, #1E3A8A 0%, #2563EB 100%);">
                            <th style="color: #FFFFFF; font-weight: 600; padding: 14px 16px; font-size: 13px; position: sticky; top: 0; z-index: 10;">Vị trí đo (POM Description)</th>
                            <th style="color: #FFFFFF; font-weight: 600; padding: 14px 16px; font-size: 13px; position: sticky; top: 0; z-index: 10;">{lbl_a}</th>
                            <th style="color: #FFFFFF; font-weight: 600; padding: 14px 16px; font-size: 13px; position: sticky; top: 0; z-index: 10;">{lbl_b}</th>
                            <th style="color: #FFFFFF; font-weight: 600; padding: 14px 16px; font-size: 13px; text-align: center; width: 150px; position: sticky; top: 0; z-index: 10;">Sai lệch (Delta)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_body_html}
                    </tbody>
                </table>
            </div>
            """
            st.markdown(full_table_render, unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            # --- KHỐI ĐỔ MÀU EXCEL ĐỒNG BỘ GIAO DIỆN ---
            df_xl = pd.DataFrame(compare_rows_for_df)
            towrite = io.BytesIO()
            with pd.ExcelWriter(towrite, engine='xlsxwriter') as writer:
                df_xl.to_excel(writer, index=False, sheet_name='Spec_Report')
                workbook  = writer.book
                worksheet = writer.sheets['Spec_Report']
                
                header_fmt = workbook.add_format({'bold':True, 'text_wrap':True, 'fg_color':'#1E3A8A', 'font_color':'white', 'border':1, 'align':'center', 'valign':'vcenter'})
                left_fmt   = workbook.add_format({'align':'left', 'valign':'vcenter', 'border':1, 'font_name':'Arial', 'font_size':10})
                center_fmt = workbook.add_format({'align':'center', 'valign':'vcenter', 'border':1, 'font_name':'Arial', 'font_size':10})
                
                green_fmt  = workbook.add_format({'bold':True, 'align':'center', 'valign':'vcenter', 'fg_color':'#E8F5E9', 'font_color':'#166534', 'border':1})
                red_fmt    = workbook.add_format({'bold':True, 'align':'center', 'valign':'vcenter', 'fg_color':'#FFEBEE', 'font_color':'#991B1B', 'border':1})
                na_fmt     = workbook.add_format({'italic':True, 'align':'center', 'valign':'vcenter', 'fg_color':'#F8FAFC', 'font_color':'#94A3B8', 'border':1})
                
                for col_num, title in enumerate(df_xl.columns):
                    worksheet.write(0, col_num, title, header_fmt)
                    max_len = max(df_xl[title].astype(str).map(len).max(), len(title)) + 4
                    worksheet.set_column(col_num, col_num, max_len)
                
                for idx, row in df_xl.iterrows():
                    worksheet.write(idx + 1, 0, row["Vị trí đo (POM)"], left_fmt)
                    worksheet.write(idx + 1, 1, row[lbl_a], center_fmt)
                    worksheet.write(idx + 1, 2, row[lbl_b], center_fmt)
                    
                    d_val = row["Sai lệch (Delta)"]
                    if d_val == "N/A":
                        worksheet.write(idx + 1, 3, "N/A", na_fmt)
                    elif d_val > 0:
                        worksheet.write(idx + 1, 3, f"+{d_val}", green_fmt)
                    elif d_val < 0:
                        worksheet.write(idx + 1, 3, d_val, red_fmt)
                    else:
                        worksheet.write(idx + 1, 3, "0.00", center_fmt)
                        
                worksheet.set_row(0, 26)
                worksheet.freeze_panes(1, 0)
                
            towrite.seek(0)
            st.download_button(label="📥 Tải Báo Cáo Đối Chiếu Có Màu (Excel)", data=towrite, file_name=f"Spec_Comparison_{style_a}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")





import io
import json
import re
import requests
import streamlit as st
import pandas as pd
from urllib.parse import quote
from google import genai
from google.genai import types

try:
    from pdf2image import convert_from_bytes, pdfinfo_from_bytes
except ImportError:
    pass

def parse_fraction(val_str):
    """
    HÀM QUY ĐỔI PHÂN SỐ NGÀNH MAY CHUẨN PPJ
    Chuyển đổi chính xác các dạng chuỗi như '1 1/2', '3/4', '1.5"' về dạng float.
    """
    if not val_str: 
        return 0.0
    val_str = str(val_str).strip().lower()
    val_str = val_str.replace('"', '').replace('inches', '').replace('inch', '').replace('s', '').strip()
    try:
        if ' ' in val_str:
            parts = [p for p in val_str.split(' ') if p.strip()]
            if len(parts) >= 2:
                whole = float(parts[0])
                frac_str = parts[1]
            else:
                whole = 0.0
                frac_str = parts[0]
        else:
            whole = 0.0
            frac_str = val_str
            
        if '/' in frac_str:
            num, denom = frac_str.split('/')
            return whole + (float(num) / float(denom))
        return float(val_str) if val_str else 0.0
    except Exception:
        return 0.0
def ai_consumption_analyst_engine(client, user_message, matched_techpack, bom_records, new_style_measurements, target_new_sketch_bytes, detected_size):
    """
    Bộ não xử lý tính toán định mức nâng cao bằng đơn vị YARD (Yds).
    Tự động kích hoạt cơ chế Vecto Hình Học Ngành May nếu KHÔNG CÓ mã tương đồng.
    Tích hợp biên may 0.44", dò tìm lai và quy tắc tách biệt Quần/Áo (Chống lộn Nẹp).
    """
    style_old_name = matched_techpack.get("StyleName", "N/A") if matched_techpack else "N/A"
    specs_old = matched_techpack.get("DetailedMeasurements", {}) if matched_techpack else {}
    
    bom_summary = ""
    if bom_records:
        bom_summary = "\n".join([f"- Vật tư: {r.get('consumption_type')}, Mã vải: {r.get('article_name')}, Khổ vải gốc: {r.get('material_size')}, ĐM gốc: {r.get('consumption_value')}" for r in bom_records])

    shrinkage_width = re.findall(r'(?:CO RÚT NGANG|NGANG)\s*(\d+(?:\.\d+)?)\s*%', user_message.upper())
    shrinkage_length = re.findall(r'(?:CO RÚT DỌC|DỌC)\s*(\d+(?:\.\d+)?)\s*%', user_message.upper())
    new_fabric_width = re.findall(r'(?:KHỔ VẢI|KHỔ)\s*(\d+)\s*(?:\"|INCH|INCHES)?', user_message.upper())

    w_shrink = float(shrinkage_width) if shrinkage_width else 0.0
    l_shrink = float(shrinkage_length) if shrinkage_length else 0.0
    f_width = float(new_fabric_width) if new_fabric_width else 0.0

    if matched_techpack:
        scenario_instruction = f"""
        KỊCH BẢN: ĐỒNG DẠNG KHO (Sử dụng dữ liệu đối chứng lịch sử)
        - Đối chiếu chênh lệch diện tích cấu trúc giữa Spec mới và Spec cũ {json.dumps(specs_old)}.
        - Bù trừ định mức tăng/giảm từ nền tảng BOM gốc: {bom_summary}.
        """
    else:
        scenario_instruction = f"""
        KỊCH BẢN: PHÂN TÍCH VECTOR HÌNH HỌC NGÀNH MAY KHÔNG CÓ MÃ TƯƠNG ĐỒNG (GEOMETRIC LAYOUT ESTIMATION)
        - Bạn không có dữ liệu lịch sử. Hãy dựa hoàn toàn vào bảng thông số 'New Spec (POM)' và hình ảnh bản vẽ rập chi tiết 'Flat Sketch' đính kèm.
        - Hãy tính diện tích hình học rập mẫu thô của từng chi tiết cấu thành dựa trên đúng phân loại sản phẩm.
        - Tính ĐM Vải chính: Cộng dồn chiều dài các mảnh rập sau khi cộng biên may, nhân hệ số hao hụt rải vải tiêu chuẩn ngành xếp trên khổ vải: {f_width if f_width > 0 else '58 inch'}.
        """

    system_instruction = f"""
    You are a strict Industrial Garment Costing Engineer at PPJ Group. 
    Your answers must mimic ChatGPT's advanced code interpreter mode:
    1. STRICT UNIT REQUIRED: All consumption values and fabric calculation results MUST be presented in YARDS (Yds) or Inches. NEVER use meters or cm.
    2. DIRECT ANSWER FIRST: Output the exact final average consumption value in YARDS (Yds) in the very first sentence.
    3. STEP-BY-STEP MATHEMATICS: Present your logic using short, punchy bullet points showing raw numbers, shrinkage multipliers, and layout area deltas.
    4. LANGUAGE: Answer directly in Vietnamese, using precise apparel terminology (co rút, định mức, hao hụt, khổ vải, nẹp liền, nẹp rời).
    
    FACTORY SEWING SEAM ALLOWANCE RULES (CRITICAL):
    - Standard Seam Allowance: ALWAYS add 0.44 inches to all general component seams (Thân trước, thân sau, sườn, giàng, dọc quần, tra cạp, v.v.).
    - Pocket Openings (Miệng túi): EXCLUDE the 0.44" rule. Pocket trims and facings must follow the exact techpack dimensions.
    - Garment Hem / Bottom Hem (Lai áo / Lai quần): DO NOT use 0.44". You MUST scan the 'New Spec (POM)' below to find the specific values for keywords like 'Hem', 'Bottom Width', 'Sleeve Hemfold'. Use that exact Techpack value for the hem calculation. If not specified, note it down.

    GARMENT CATEGORY SPECIFIC RULES (STRICT SEPARATION TO AVOID ERROR):
    
    1. IF THE CATEGORY IS PANT / SHORT / SKORT / TROUSER (NHÓM HÀNG QUẦN):
       - STICK TO JEANS/PANTS LOGIC ONLY.
       - ABSOLUTELY FORBIDDEN: Do NOT apply Shirt Placket rules (Cấm áp dụng quy tắc nẹp rời/nẹp liền gập cuốn của áo). 
       - Waistband Construction (Cạp quần): Calculate waistband fabric based strictly on Waistband Height and Waist Circumference.
       - Zipper Fly / Fly Placket (Cửa quần): Only calculate based on standard front fly extensions (typically adding a small standard extension of 1.5" to 2" for the fly j-stitch width on ONE side of the left front panel only. Do NOT multiply by 2 across both front panels like a shirt).
       - Coin Pocket (Túi đồng xu): Check for coin pocket width/height if specified.
       
    2. IF THE CATEGORY IS SHIRT / JACKET / TOP / COAT (NHÓM HÀNG ÁO):
       - NẸP LIỀN (Fold-on/Extended Placket): If folded from the main body, add 2 times the placket width extension to the raw width of each front body panel pattern before calculating fabric consumption.
       - NẸP RỜI (Separate Placket): Treat it as a standalone independent geometric pattern strip panel (Length = body length + seams, Width = placket width x 2 + standard seams 2 x 0.44"). Add this separate piece to the overall layout marker area.

    {scenario_instruction}

    CRITICAL DATA FOR CALCULATION:
    1. NEW STYLE TECHPACK DATA:
       - Target Base Size detected: Size {detected_size}
       - New Spec (POM) parsed by vision: {json.dumps(new_style_measurements)}
    2. USER INPUT FABRIC CHANGES:
       - Fabric Width requested: {f_width if f_width > 0 else '58 inch'}
       - Width Shrinkage (Co rút ngang): {w_shrink}%
       - Length Shrinkage (Co rút dọc): {l_shrink}%
    """

    chat_contents = [types.Part.from_text(text=system_instruction)]
    for past_chat in st.session_state.get("consumption_chat_history", []):
        chat_contents.append(types.Part.from_text(text=f"User: {past_chat['user']}"))
        chat_contents.append(types.Part.from_text(text=f"AI: {past_chat['ai']}"))
        
    chat_contents.append(types.Part.from_text(text=f"User current request: {user_message}"))
    if target_new_sketch_bytes:
        chat_contents.append(types.Part.from_bytes(data=target_new_sketch_bytes, mime_type='image/jpeg'))

    try:
        response = client.models.generate_content(model='gemini-2.5-flash', contents=chat_contents)
        ai_reply = response.text if response.text else "Hệ thống AI không thể đưa ra phân tích."
        st.session_state["consumption_chat_history"].append({"user": user_message, "ai": ai_reply})
        return ai_reply
    except Exception as e:
        return f"🚨 Lỗi cổng phân tích định mức: {str(e)}"

if "get_secure_gemini_key" in globals():
    gemini_key = get_secure_gemini_key()
else:
    gemini_key = st.secrets.get("GEMINI_API_KEY", "").strip()

if gemini_key:
    client = genai.Client(api_key=gemini_key, http_options=types.HttpOptions(api_version='v1'))

def process_single_pdf_batch(file_bytes, file_name):
    """
    Hàm bóc tách dữ liệu kỹ thuật từ một file PDF độc lập.
    Quét đặc biệt thông số Lai và chi tiết Nẹp áo để phục vụ luồng tính toán chừa biên may.
    """
    import time
    try:
        if "get_secure_gemini_key" in globals():
            gemini_key_local = get_secure_gemini_key()
        else:
            gemini_key_local = st.secrets.get("GEMINI_API_KEY", "").strip()
            
        if not gemini_key_local:
            return {"success": False, "error": "API Key cho Gemini đang bị thiếu trong Secrets."}
            
        client_ai = genai.Client(api_key=gemini_key_local)
        info = pdfinfo_from_bytes(file_bytes)
        total_p = int(info.get("Pages", 1))
        
        pdf_parts_payload = []
        chat_images = convert_from_bytes(file_bytes, dpi=90, first_page=1, last_page=total_p)
        
        stored_pages_bytes = []
        for page_img in chat_images:
            img_buf = io.BytesIO()
            page_img.convert("RGB").save(img_buf, format="JPEG", quality=75)
            img_data = img_buf.getvalue()
            stored_pages_bytes.append(img_data)
            pdf_parts_payload.append(types.Part.from_bytes(data=img_data, mime_type='image/jpeg'))
            
        industrial_extraction_prompt = (
            "You are an expert Garment Specification Auditor at PPJ Group. Analyze all attached sheets page by page. "
            "1. Identify the core 'Base Size' / 'Sample Size'. "
            "2. Identify the Buyer name and Category (Pant/Shirt/Jacket). "
            "3. Find the exact 'Style ID' / 'Style Number'. "
            "4. Extract the entire grading matrix table columns for ALL available sizes. "
            "5. Find the exact PAGE INDEX (0-based) that contains the FULL BODY APPAREL FLAT SKETCH. "
            "6. CRITICAL APPRAISAL FOR HEM & PLACKET DETAILS: Pay extreme attention to bottom hem allowances. If the category is a Shirt or Jacket, scan for 'Placket Width', 'Center Front Placket', or center stitching lines. Identify if the placket is separate or grown-on/folded, and record its measurement inside the measurements dictionary accurately. "
            "Return a completely valid raw JSON string matching this schema (no markdown blocks): "
            "{"
            "  \"style_number_parsed\": \"string\","
            "  \"buyer\": \"string\","
            "  \"category\": \"string\","
            "  \"base_size_name\": \"string\","
            "  \"sketch_page_index_detected\": 0,"
            "  \"measurements\": {\"POM Description\": \"Value\"},"
            "  \"full_size_matrix\": {\"POM Description\": {\"Size_Name\": \"Value\"}}"
            "}"
        )
        pdf_parts_payload.append(types.Part.from_text(text=industrial_extraction_prompt))
        
        for attempt in range(3):
            try:
                response = client_ai.models.generate_content(
                    model='gemini-2.5-flash', 
                    contents=pdf_parts_payload,
                    config={"response_mime_type": "application/json"}
                )
                if response and response.text:
                    parsed_json = json.loads(response.text)
                    sketch_idx = parsed_json.get("sketch_page_index_detected", 0)
                    extracted_sketch_bytes = stored_pages_bytes[sketch_idx] if sketch_idx < len(stored_pages_bytes) else stored_pages_bytes
                    
                    return {
                        "success": True, 
                        "data": parsed_json,
                        "sketch_bytes": extracted_sketch_bytes
                    }
            except Exception:
                time.sleep(1.5)
                continue
        return {"success": False, "error": "AI không thể cấu trúc dữ liệu JSON sau 3 lần thử."}
    except Exception as e:
        return {"success": False, "error": str(e)}
new_style_id_detected = "UNKNOWN_STYLE"
new_style_category_detected = ""
new_style_fabric_detected = "UNKNOWN_FABRIC"
new_style_measurements_dict = {}
new_style_base_size = "32"
target_new_sketch_bytes = None 

target_file_object = None
if 'uploaded_file' in st.session_state and st.session_state['uploaded_file'] is not None:
    target_file_object = st.session_state['uploaded_file']
elif 'chat_uploader' in st.session_state and st.session_state['chat_uploader'] is not None:
    target_file_object = st.session_state['chat_uploader']
elif 'bom_matrix_uploader' in st.session_state and st.session_state['bom_matrix_uploader'] is not None:
    target_file_object = st.session_state['bom_matrix_uploader']

has_file = target_file_object is not None

if has_file:
    file_bytes = target_file_object.getvalue()
    file_name = target_file_object.name
    if file_name.lower().endswith('.pdf'):
        try:
            res_pdf = process_single_pdf_batch(file_bytes, file_name)
            if res_pdf.get("success"):
                meta_p = res_pdf["data"]
                new_style_id_detected = meta_p.get("style_number_parsed", "UNKNOWN_STYLE")
                new_style_category_detected = meta_p.get("category", "")
                new_style_base_size = meta_p.get("base_size_name", "32")
                new_style_measurements_dict = meta_p.get("measurements", {})
                target_new_sketch_bytes = res_pdf.get("sketch_bytes")
        except Exception:
            pass
    else:
        target_new_sketch_bytes = file_bytes

dynamic_keyword = str(new_style_id_detected).strip().upper()
base_sb_url = SB_URL.rstrip('/')
headers = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}

if menu_selection == "🧵 BOM & Consumption Matrix":
    st.markdown('<div class="component-title-box">🧵 INTELLIGENT BOM & CONSUMPTION MATRIX ENGINE</div>', unsafe_allow_html=True)
    
    if "matched_techpack" not in st.session_state: st.session_state["matched_techpack"] = None
    if "bom_records" not in st.session_state: st.session_state["bom_records"] = []
    if "consumption_chat_history" not in st.session_state: st.session_state["consumption_chat_history"] = []

    control_col1, control_col2 = st.columns([3.3, 0.7])
    with control_col1:
        st.markdown("<p style='font-weight:700; font-size:12px; color:#1E293B;'>📁 INGEST NEW STYLE REPRINTS (PDF/IMAGE)</p>", unsafe_allow_html=True)
        st.file_uploader("Upload Techpack file", type=["pdf", "jpg", "jpeg", "png"], key="bom_matrix_uploader", label_visibility="collapsed")
            
    with control_col2:
        st.markdown("<p style='font-weight:700; font-size:12px; color:#1E293B;'>🧹 RESET CORE</p>", unsafe_allow_html=True)
        if st.button("🗑️ PURGE CHAT CACHE", key="purge_cache_matrix_btn", use_container_width=True, type="secondary"):
            st.session_state["consumption_chat_history"] = []
            st.session_state["matched_techpack"] = None
            st.session_state["bom_records"] = []
            st.success("♻️ MEMORY PURGED - SẴN SÀNG CHO MÃ HÀNG MỚI")
            st.rerun()

    st.markdown("---")

    if not has_file:
        st.info("👋 Vui lòng tải lên tệp Techpack hồ sơ thiết kế (PDF) ở phía trên để hệ thống bắt đầu quét và lập lịch trình đối soát.")
        st.stop()

    if new_style_base_size and new_style_base_size != "32":
        st.info(f"📋 **CƠ SỞ ĐỐI SOÁT KIỂM TRA:** Mẫu mới số hóa mã hàng `{new_style_id_detected}` | Quy chuẩn kích thước hình học rập mẫu: **SIZE {new_style_base_size}**")
    else:
        st.info(f"📋 **CƠ SỞ ĐỐI SOÁT KIỂM TRA:** Đang áp dụng quy chuẩn kích thước hình học rập mẫu cơ sở: **SIZE 32 / M (Mặc định)**")

    with st.spinner("🧠 Hệ thống thị giác máy tính đang quét kết cấu phom dáng Flat Sketch..."):
        try:
            headers_db = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}
            url_db = f"{base_sb_url}/rest/v1/thong_so_techpack"
            query_params = {"select": "StyleName,Buyer,Category,BaseSize,DetailedMeasurements,SketchURL,sketch_vector", "limit": 100}
            
            db_res = requests.get(url_db, headers=headers_db, params=query_params, timeout=15)
            all_historical_styles = db_res.json() if db_res.status_code == 200 else []
            
            if all_historical_styles:
                styles_pool_summary = []
                for idx, s in enumerate(all_historical_styles):
                    styles_pool_summary.append({
                        "pool_index": idx,
                        "style_name": s.get("StyleName"),
                        "sketch_features_vector": s.get("sketch_vector", "")
                    })
                
                match_prompt = f"""
                You are an expert Computer Vision Ingestion System specialized in Apparel Manufacturing at PPJ Group.
                Analyze the ATTACHED NEW SKETCH IMAGE and find the single closest matching historical garment style from the pool.
                
                STRICT APPAREL AUDIT RULES (IGNORE THE OUTLINE, FOCUS SOLELY ON INTERNAL CONSTRUCTION DETAILS):
                1. WAISTBAND CONSTRUCTION (CẠP QUẦN): Look closely at the waistband stitches. 
                   - If the new image shows a full elastic waistband (cạp chun co giãn toàn bộ, nhiều đường nhăn song song, NO belt loops, NO zipper fly), you MUST REJECT any historical styles that have belt loops (đai quần), front zipper fly (khóa kéo), or standard button closures.
                2. POCKET TYPES & PLACEMENT (KẾT CẤU TÚI): Look at the internal pocket placement.
                   - Distinguish clearly between Patch Pockets (túi ốp nổi lớn bên ngoài, túi hộp kiểu Cargo) and Slit/In-seam Pockets (túi mổ sườn phẳng). Do NOT match a clean plain front leg with a Cargo/Utility pocket design.
                3. SEAM LINES & PANEL CUTS: Look inside the body lines.
                
                HISTORICAL POOL DATA:
                {json.dumps(styles_pool_summary)}
                
                Select the single index that represents the exact match in internal technical stitching construction, NOT just the shorts frame.
                Return a raw valid JSON object inside your response: {{"selected_pool_index": 0}}
                """
                
                match_contents = [types.Part.from_text(text=match_prompt)]
                if target_new_sketch_bytes:
                    match_contents.append(types.Part.from_bytes(data=target_new_sketch_bytes, mime_type='image/jpeg'))
                    
                res_match = client.models.generate_content(model='gemini-2.5-flash', contents=match_contents)
                ai_raw_text = res_match.text.strip()
                
                json_block_clean = ""
                match_json_obj = re.search(r'\{\s*"selected_pool_index"\s*:\s*\d+\s*\}', ai_raw_text)
                if match_json_obj:
                    json_block_clean = match_json_obj.group(0).strip()
                
                if json_block_clean:
                    match_result = json.loads(json_block_clean)
                    best_idx = match_result.get("selected_pool_index", -1)
                    if 0 <= best_idx < len(all_historical_styles):
                        st.session_state["matched_techpack"] = all_historical_styles[best_idx]
        except Exception as match_err:
            st.sidebar.error(f"Lỗi hệ thống đối soát hình ảnh: {str(match_err)}")
if menu_selection == "🧵 BOM & Consumption Matrix":
    if st.session_state.get("matched_techpack"):
        try:
            target_style_name_bom = str(st.session_state["matched_techpack"].get("StyleName", "")).strip()
            url_bom = f"{SB_URL.rstrip('/')}/rest/v1/san_pham"
            
            query_bom = {
                "select": "style_name,article_name,consumption_type,material_size,uom,consumption_value,notes",
                "style_name": f"eq.{target_style_name_bom}"
            }
            res_bom = requests.get(url_bom, headers=headers, params=query_bom, timeout=15)
            if res_bom.status_code == 200 and len(res_bom.json()) > 0:
                st.session_state["bom_records"] = res_bom.json()
            else:
                core_digits = re.findall(r'\d+', target_style_name_bom)
                search_digits = max(core_digits, key=len) if core_digits else target_style_name_bom
                
                query_bom_fb = {
                    "select": "style_name,article_name,consumption_type,material_size,uom,consumption_value,notes",
                    "style_name": f"ilike.*{search_digits}*"
                }
                res_bom_fb = requests.get(url_bom, headers=headers, params=query_bom_fb, timeout=15)
                if res_bom_fb.status_code == 200:
                    raw_list = res_bom_fb.json()
                    st.session_state["bom_records"] = [r for r in raw_list if search_digits in str(r.get("style_name", ""))]
                else:
                    st.session_state["bom_records"] = []
        except Exception:
            st.session_state["bom_records"] = []

    matched_techpack = st.session_state.get("matched_techpack")
    bom_records = st.session_state.get("bom_records", [])

    st.markdown("### 🖼️ ĐỐI CHIẾU SỰ TƯƠNG ĐỒNG HÌNH ẢNH THIẾT KẾ (FLAT SKETCH)")
    img_col1, img_col2 = st.columns(2)
    with img_col1:
        if target_new_sketch_bytes is not None:
            st.image(target_new_sketch_bytes, caption=f"Mẫu mới tải lên ({new_style_id_detected})", use_container_width=True)

    with img_col2:
        if matched_techpack:
            target_style_name = str(matched_techpack.get("StyleName", "Mẫu tương đồng")).strip().upper()
            st.markdown(f"<p style='color: #1E3A8A; font-size: 13px; font-weight: 700; margin-bottom: 8px; text-align: center;'>🎯 Mã tương đồng trong kho: {target_style_name}</p>", unsafe_allow_html=True)
            
            auth_headers = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}
            url_options = [
                f"{SB_URL.rstrip('/')}/storage/v1/object/public/kho_anh/{target_style_name}.jpg",
                f"{SB_URL.rstrip('/')}/storage/v1/object/public/kho_anh/{target_style_name}.JPG",
                f"{SB_URL.rstrip('/')}/storage/v1/object/public/kho_anh/{target_style_name}.jpeg",
                f"{SB_URL.rstrip('/')}/storage/v1/object/public/kho_anh/{target_style_name.lower()}.jpg"
            ]
            
            img_content_final = None
            for url_opt in url_options:
                try:
                    img_response = requests.get(url_opt, headers=auth_headers, timeout=8)
                    if img_response.status_code == 200 and len(img_response.content) > 500:
                        if img_response.content.startswith(b'\xff\xd8') or b'<!DOCTYPE' not in img_response.content[:100]:
                            img_content_final = img_response.content
                            break
                except Exception:
                    continue
                    
            if img_content_final:
                try:
                    st.image(img_content_final, caption=f"Ảnh bản vẽ gốc của mã {target_style_name}", use_container_width=True)
                except Exception:
                    st.warning("⚠️ Không thể kết xuất tệp đồ họa do lỗi định dạng nhị phân vật lý.")
            else:
                db_stored_url = matched_techpack.get("SketchURL") or matched_techpack.get("sketch_url", "")
                if db_stored_url and "public/kho_anh" not in str(db_stored_url):
                    try:
                        st.image(db_stored_url, caption=f"Ảnh bản vẽ gốc mã {target_style_name} (Direct Link)", use_container_width=True)
                    except Exception:
                        st.info("⚠️ Không có ảnh vật lý nào khớp trên Cloud Storage.")
                else:
                    st.info("⚠️ Không có ảnh vật lý nào khớp trên Cloud Storage.")
        else:
            st.warning("⚠️ KHÔNG TÌM THẤY MÃ TƯƠNG ĐỒNG TRONG KHO! Tự động kích hoạt cơ chế tính toán độc lập bằng Vector Hình Học Ngành May.")

    st.markdown("<br>### 📐 SO SÁNH HAI BẢNG THÔNG SỐ KỸ THUẬT RẬP MẪU", unsafe_allow_html=True)
    spec_col1, spec_col2 = st.columns(2)

    with spec_col1:
        st.markdown(f"📊 **Bảng 1: Thông số Mẫu mới nạp ({new_style_base_size})**")
        if new_style_measurements_dict:
            df_new_spec = pd.DataFrame(list(new_style_measurements_dict.items()), columns=["Vị trí đo (POM Description)", "Thông số mới"])
        else:
            df_new_spec = pd.DataFrame(columns=["Vị trí đo (POM Description)", "Thông số mới"])
        st.dataframe(df_new_spec, use_container_width=True, hide_index=True)
        
    with spec_col2:
        if matched_techpack:
            old_style_title = str(matched_techpack.get("StyleName", "N/A")).upper()
            old_size_title = matched_techpack.get("BaseSize", "N/A")
            st.markdown(f"📋 **Bảng 2: Thông số Mã trong kho ({old_style_title}) [SIZE {old_size_title}]**")
            old_specs = matched_techpack.get("DetailedMeasurements", {})
            if old_specs:
                df_old_spec = pd.DataFrame(list(old_specs.items()), columns=["Vị trí đo (POM Description)", "Thông số cũ"])
            else:
                df_old_spec = pd.DataFrame(columns=["Vị trí đo (POM Description)", "Thông số cũ"])
            st.dataframe(df_old_spec, use_container_width=True, hide_index=True)
        else:
            st.markdown("📋 **Bảng 2: Thông số Mã tương đồng trong kho**")
            st.info("💡 Trạng thái: Trống dữ liệu kho. Hệ thống sẵn sàng tính toán diện tích rập mô phỏng tự động.")

    if matched_techpack and bom_records:
        st.markdown("<br>📦 **Chi Tiết Định Mức Định Hình (BOM Lịch Sử của Mã hàng cũ):**", unsafe_allow_html=True)
        formatted_bom = []
        for r in bom_records:
            def clean_nan(v): return "" if (not v or str(v).lower() in ["nan", "none", "null"]) else str(v).strip()
            formatted_bom.append({
                "Mã hàng đối chứng": clean_nan(r.get("style_name")).upper(),
                "Loại nguyên vật liệu": clean_nan(r.get("consumption_type")),
                "Chi tiết vật tư (Article)": clean_nan(r.get("article_name")),
                "Khổ / Cỡ vật tư": clean_nan(r.get("material_size")),
                "Định mức gốc": clean_nan(r.get("consumption_value")),
                "UOM": clean_nan(r.get("uom"))
            })
        st.dataframe(pd.DataFrame(formatted_bom), use_container_width=True, hide_index=True)

    st.markdown("<br><hr style='border:0.5px solid #CBD5E1;'>", unsafe_allow_html=True)
    
           # THIẾT KẾ CỤM ĐIỀU KHIỂN CHAT BOX THÔNG MINH SIÊU TỐC
    chat_header_col1, chat_header_col2 = st.columns([3.2, 0.8])
    with chat_header_col1:
        st.markdown("### 💬 TRỢ LÝ AI PHÂN TÍCH ĐỊNH MỨC SẢN XUẤT (HỎI ĐÂU ĐÁP ĐÓ)")
    with chat_header_col2:
        # ✅ SỬA LỖI SIÊU TỐC: Xóa trực tiếp mảng và không ép hệ thống phải load lại file PDF từ đầu
        if st.button("🗑️ XÓA LỊCH SỬ CHAT", key="direct_clear_chat_btn", use_container_width=True):
            st.session_state["consumption_chat_history"] = []
            st.toast("♻️ Đã xóa sạch lịch sử chat tức thì!")

    chat_container = st.container()
    with chat_container:
        for chat in st.session_state.get("consumption_chat_history", []):
            with st.chat_message("user"): 
                st.write(chat["user"])
            with st.chat_message("assistant"): 
                st.write(chat["ai"])
                
    if user_query := st.chat_input("Nhập yêu cầu phân tích (Ví dụ: Tính định mức vải chính khi co rút ngang 5%, dọc 3%)..."):
        with chat_container:
            with st.chat_message("user"):
                st.write(user_query)
                
            with st.chat_message("assistant"):
                with st.spinner("🤖 AI đang phân tích dữ liệu và tính toán định mức..."):
                    ai_reply = ai_consumption_analyst_engine(
                        client=client,
                        user_message=user_query,
                        matched_techpack=matched_techpack,
                        bom_records=bom_records,
                        new_style_measurements=new_style_measurements_dict,
                        target_new_sketch_bytes=target_new_sketch_bytes,
                        detected_size=new_style_base_size
                    )
                    st.write(ai_reply)
        
        # ✅ THUẬT TOÁN ĐÓNG ĐINH NEO CUỘN: Ép trình duyệt tự động scroll lướt màn hình xuống vị trí tin nhắn cuối cùng
        st.components.v1.html(
            """
            <script>
                var doc = window.parent.document;
                var sections = doc.querySelectorAll('section.main');
                if (sections.length > 0) {
                    sections[0].scrollTo({top: sections[0].scrollHeight, behavior: 'smooth'});
                }
            </script>
            """,
            height=0,
        )








# -----------------------------------------------------------------------------
# CHỨC NĂNG 3: QUẢN LÝ ĐỊNH MỨC MUA SẮM VÀ ĐẶT HÀNG (PURCHASE CONSUMPTION)
# -----------------------------------------------------------------------------
elif menu_selection == "🛒 Purchase Consumption":
    st.markdown('<div class="component-title-box">🛒 PURCHASE CONSUMPTION & INTELLIGENT PLANNING ENGINE</div>', unsafe_allow_html=True)
    
    # =============================================================================
    # ĐOẠN 1: KHỞI TẠO BỘ NHỚ STATE ĐA NĂNG (HỖ TRỢ CẢ KIỂU DANH SÁCH VÀ TỪ ĐIỂN)
    # =============================================================================
    if "purchase_ready" not in st.session_state: st.session_state["purchase_ready"] = False
    if "sbd_parsed_data" not in st.session_state: st.session_state["sbd_parsed_data"] = {}
    
    # Khởi tạo đa cấu trúc dự phòng để không làm gãy code hiển thị cũ phía dưới của bạn
    if "pur_tp_parsed_data" not in st.session_state: st.session_state["pur_tp_parsed_data"] = {"success": False, "data": []}
    
    if "step1_marker_ready" not in st.session_state: st.session_state["step1_marker_ready"] = False
    if "step2_computation_active" not in st.session_state: st.session_state["step2_computation_active"] = False
    if "bulk_cad_data_store" not in st.session_state: st.session_state["bulk_cad_data_store"] = []

    menu_sub = st.radio(
        "💡 CHỌN CÔNG ĐOẠN TÁC NGHIỆP THỰC HIỆN:",
        ["🧠 CHỨC NĂNG 1: TRỢ LÝ AI TÍNH ĐỊNH MỨC TRUNG BÌNH (CẦN SBD + TECHPACK)", 
         "✂️ CHỨC NĂNG 2: MÁY TÍNH TÁC NGHIỆP BÀN CẮT TỰ ĐỘNG & LƯU KHO (CHỈ CẦN FILE SBD)"],
        horizontal=True, key="purchase_sub_menu_root"
    )
    
    st.markdown("---")

    # (Giữ nguyên đoạn code tra cứu kho lưu trữ lịch sử của Chức năng 2 ở đây nếu có...)

        # =============================================================================
    # ĐOẠN A: TIẾP NHẬN FILE VÀ XỬ LÝ SBD BƯỚC 1 (Đã chống lỗi Pydantic)
    # =============================================================================
    if menu_sub.startswith("🧠 CHỨC NĂNG 1"):
        col_left, col_right = st.columns(2)
        with col_left: 
            file_sbd = st.file_uploader("📋 Chọn File SBD Số Lượng (Excel/PDF)", type=["xlsx", "xls", "pdf"], key="purchase_sbd_c1")
        with col_right: 
            file_tp = st.file_uploader("📐 Chọn File Techpack Thông Số (PDF)", type=["pdf"], key="purchase_tp_c1")
            
            if file_tp:
                col_p1, col_p2 = st.columns(2)
                with col_p1: 
                    start_page = st.number_input("🔹 Từ trang", min_value=1, value=5, step=1, key="tp_start_p")
                with col_p2: 
                    end_page = st.number_input("🔸 Đến trang", min_value=1, value=7, step=1, key="tp_end_p")
        
        if file_sbd and file_tp:
            trigger_btn = st.button("⚡ KÍCH HOẠT SỐ HÓA ĐA LUỒNG SONG SONG", type="primary", use_container_width=True, key="activate_parallel_ingest_c1")
            if trigger_btn:
                st.cache_data.clear()
                
                if "get_secure_gemini_key" in globals(): gemini_key = get_secure_gemini_key()
                else: gemini_key = st.secrets.get("GEMINI_API_KEY", "").strip()
                
                if not gemini_key:
                    st.error("❌ Không tìm thấy API KEY trong cấu hình hệ thống (Secrets)!")
                    st.stop()
                
                import json
                import io
                
                # Tự động nhận diện phiên bản thư viện cài trên server để chọn cú pháp đúng
                is_new_sdk = False
                try:
                    from google import genai
                    from google.genai import types
                    client_ai = genai.Client(api_key=gemini_key)
                    is_new_sdk = True
                except Exception:
                    try:
                        import google.generativeai as google_genai
                        google_genai.configure(api_key=gemini_key)
                    except Exception as err_import:
                        st.error(f"❌ Lỗi thư viện Gemini: {str(err_import)}")
                        st.stop()

                # --- TIẾN HÀNH XỬ LÝ BƯỚC 1/2: SBD ---
                with st.spinner("🧠 Bước 1/2: AI đang bóc tách File SBD số lượng..."):
                    sbd_bytes = file_sbd.getvalue()
                    sbd_content_str = ""
                    
                    sbd_prompt = """You are a garment production AI. Analyze the 'Quantity Details' table inside this garment order sheet.
                    CRITICAL TABLE STRUCTURE INSTRUCTIONS:
                    1. The size header is split into TWO stacked rows. For example, a column has '26 X' on the upper row and '30' directly below it on the next row. You must combine them vertically to form the full size name, e.g., '26 X 30'.
                    2. Underneath this combined size row, extract the actual production quantity numbers (e.g., 88, 156, 150, 122...).
                    3. Identify the 'Style' or 'Key Item' number.
                    4. Find the 'Total' quantity.

                    Strictly output a clean, valid raw JSON object matching this schema exactly, do not include any markdown format tags:
                    {
                      "style_id": "string",
                      "total_quantity": integer,
                      "size_breakdown": {
                        "Size Name": integer
                      }
                    }"""
                    
                    if file_sbd.name.lower().endswith(('.xlsx', '.xls')):
                        try:
                            import pandas as pd
                            excel_data = pd.read_excel(io.BytesIO(sbd_bytes), sheet_name=None)
                            for sheet_name, df_sheet in excel_data.items():
                                sbd_content_str += f"\n--- SHEET: {sheet_name} ---\n{df_sheet.fillna('').to_csv(index=False)}"
                        except Exception as e:
                            st.error(f"❌ Lỗi đọc file Excel SBD: {str(e)}")
                            st.stop()

                    try:
                        if is_new_sdk:
                            sbd_parts = []
                            if sbd_content_str:
                                sbd_parts.append(types.Part.from_text(text=sbd_content_str))
                            else:
                                sbd_parts.append(types.Part.from_bytes(data=sbd_bytes, mime_type="application/pdf"))
                            sbd_parts.append(types.Part.from_text(text=sbd_prompt))
                            
                            res_sbd = client_ai.models.generate_content(
                                model='gemini-2.5-flash', 
                                contents=sbd_parts, 
                                config=types.GenerateContentConfig(response_mime_type="application/json")
                            )
                            raw_text_sbd = res_sbd.text
                        else:
                            model_old = google_genai.GenerativeModel('gemini-2.5-flash')
                            sbd_parts = []
                            if sbd_content_str:
                                sbd_parts.append(sbd_content_str)
                            else:
                                sbd_parts.append({"mime_type": "application/pdf", "data": sbd_bytes})
                            sbd_parts.append(sbd_prompt)
                            
                            res_sbd = model_old.generate_content(sbd_parts, generation_config={"response_mime_type": "application/json"})
                            raw_text_sbd = res_sbd.text

                        clean_text_sbd = raw_text_sbd.strip().replace("```json", "").replace("```", "").strip()
                        st.session_state["sbd_parsed_data"] = json.loads(clean_text_sbd)
                    except Exception as e:
                        st.error(f"❌ Gemini gặp sự cố tại Bước 1 (SBD): {str(e)}")
                        st.stop()
                               # =============================================================================
                # ĐOẠN B: SỐ HÓA THÔNG SỐ TECHPACK BƯỚC 2 (CẮT TRANG RAM VẬT LÝ BẰNG PYPDF2)
                # =============================================================================
                with st.spinner(f"📐 Bước 2/2: Đang dùng AI bóc tách thông số Techpack (Trang {start_page} đến {end_page})..."):
                    try:
                        file_tp_bytes = file_tp.getvalue()
                        trimmed_pdf_bytes = file_tp_bytes
                        
                        # Thực hiện cắt nhỏ file PDF ngay trên RAM bằng PyPDF2 để giảm tải cho AI
                        try:
                            import PyPDF2
                            original_pdf = PyPDF2.PdfReader(io.BytesIO(file_tp_bytes))
                            pdf_writer = PyPDF2.PdfWriter()
                            
                            # Vòng lặp trích xuất đúng dải trang người dùng cấu hình (chuyển hệ số từ 1 về 0-index)
                            for page_num in range(start_page - 1, min(end_page, len(original_pdf.pages))):
                                pdf_writer.add_page(original_pdf.pages[page_num])
                                
                            trimmed_buffer = io.BytesIO()
                            pdf_writer.write(trimmed_buffer)
                            trimmed_pdf_bytes = trimmed_buffer.getvalue()
                        except Exception:
                            pass # Nếu server thiếu thư viện, hệ thống tự động fallback gửi file gốc an toàn
                        
                        tp_prompt = """You are an expert garment patternmaker and data analyst.
                        Analyze the Size Specification or Measurement Chart table on the provided document pages.
                        
                        CRITICAL CONVERSION RULES FOR GARMENT FRACTIONS:
                        1. Convert ALL fractional measurements into clean decimal numbers before writing to JSON (e.g., 15 ½ to 15.5, 14 ¼ to 14.25).
                        2. If a cell contains a pure integer with no fraction, keep it as an integer (e.g., '16' remains 16).
                        
                        Extract all measurement points (POM), their descriptions, and the values for each size column.
                        
                        Strictly return a clean, valid raw JSON object wrapping the measurement data:
                        {
                          "status": "success",
                          "success": true,
                          "data": [
                            {
                              "pom_id": "string or null",
                              "description": "string",
                              "measurements": {
                                "Size Name": "string or number"
                              }
                            }
                          ]
                        }
                        Do not include any markdown syntax wrapping."""

                        # Gọi lệnh API Techpack dựa theo thư viện đang chạy thực tế trên server
                        if is_new_sdk:
                            tp_parts = [
                                types.Part.from_bytes(data=trimmed_pdf_bytes, mime_type="application/pdf"),
                                types.Part.from_text(text=tp_prompt)
                            ]
                            res_tp_ai = client_ai.models.generate_content(
                                model='gemini-2.5-flash', 
                                contents=tp_parts, 
                                config=types.GenerateContentConfig(response_mime_type="application/json")
                            )
                            raw_text_tp = res_tp_ai.text
                        else:
                            model_old = google_genai.GenerativeModel('gemini-2.5-flash')
                            tp_parts = [
                                {"mime_type": "application/pdf", "data": trimmed_pdf_bytes},
                                tp_prompt
                            ]
                            res_tp_ai = model_old.generate_content(tp_parts, generation_config={"response_mime_type": "application/json"})
                            raw_text_tp = res_tp_ai.text

                        clean_text_tp = raw_text_tp.strip().replace("```json", "").replace("```", "").strip()
                        parsed_json = json.loads(clean_text_tp)
                        
                        # Trích xuất danh sách mảng dữ liệu phẳng từ kết quả trả về của AI
                        extracted_list = parsed_json.get("data", parsed_json.get("spec_table", []))
                        
                        # Đổ dữ liệu đồng nhất vào mọi định dạng cấu trúc mà bộ hiển thị phía dưới của bạn có thể gọi
                        st.session_state["pur_tp_parsed_data"] = {
                            "success": True,
                            "data": extracted_list,
                            "spec_table": extracted_list
                        }
                        
                    except Exception as e:
                        st.error(f"❌ Gemini gặp sự cố tại Bước 2 (Techpack): {str(e)}")
                        st.stop()
                    
                # Hoàn thành xuất sắc cả 2 phân hệ, kích hoạt và tải lại trang để render kết quả
                st.session_state["purchase_ready"] = True
                st.rerun()
                # Hoàn thành bóc tách xuất sắc cả 2 phân hệ
                st.session_state["purchase_ready"] = True
                st.rerun()

    # =============================================================================
    # ĐOẠN BỔ SUNG: TỰ ĐỘNG HIỂN THỊ KẾT QUẢ SỐ HÓA RA MÀN HÌNH SAU KHI RERUN
    # =============================================================================
    if st.session_state.get("purchase_ready") and menu_sub.startswith("🧠 CHỨC NĂNG 1"):
        st.markdown("<p style='font-weight:700; font-size:16px; color:#10B981; margin-top:15px;'>🎉 KẾT QUẢ AI SỐ HÓA DỮ LIỆU THÀNH CÔNG</p>", unsafe_allow_html=True)
        
        col_view1, col_view2 = st.columns(2)
        
        with col_view1:
            st.markdown("##### 📋 Ma Trận Sản Lượng Đơn Hàng (SBD)")
            sbd_res = st.session_state.get("sbd_parsed_data", {})
            if sbd_res:
                st.write(f"**Mã sản phẩm (Style ID):** `{sbd_res.get('style_id', 'N/A')}`")
                st.write(f"**Tổng sản lượng đơn:** {sbd_res.get('total_quantity', 0):,}")
                
                breakdown = sbd_res.get("size_breakdown", {})
                if breakdown:
                    import pandas as pd
                    df_sbd_table = pd.DataFrame(list(breakdown.items()), columns=["Kích cỡ (Size)", "Số lượng"])
                    st.dataframe(df_sbd_table, use_container_width=True, hide_index=True)
            else:
                st.info("Chưa có dữ liệu phân bổ sản lượng SBD.")
                
        with col_view2:
            st.markdown("##### 📐 Bảng Thông Số Bản Vẽ Kỹ Thuật (Techpack)")
            tp_res = st.session_state.get("pur_tp_parsed_data", {})
            
            # Giải mã linh hoạt cấu trúc bọc dữ liệu đa tầng để ép hiển thị ra màn hình
            raw_list = []
            if isinstance(tp_res, list):
                raw_list = tp_res
            elif isinstance(tp_res, dict):
                raw_list = tp_res.get("data", tp_res.get("spec_table", []))
                
            if raw_list:
                import pandas as pd
                flat_rows = []
                for item in raw_list:
                    pom = item.get("pom_id", "") or ""
                    desc = item.get("description", "")
                    sizes_val = item.get("measurements", {})
                    
                    # Ghép phẳng toàn bộ thông tin mã POM, mô tả và dải size thành 1 dòng dữ liệu
                    row_data = {"Mã POM": pom, "Mô tả chi tiết": desc}
                    row_data.update(sizes_val)
                    flat_rows.append(row_data)
                    
                df_tp_table = pd.DataFrame(flat_rows)
                st.dataframe(df_tp_table, use_container_width=True, hide_index=True)
            else:
                st.info("Chưa có dữ liệu thông số Techpack.")
                
        st.markdown("<hr style='border:1px solid #E2E8F0; margin-top:20px; margin-bottom:20px;'>", unsafe_allow_html=True)
                # =============================================================================
               # =============================================================================
               # =============================================================================
                # =============================================================================
                # =============================================================================
               # =============================================================================
        # ĐOẠN 1: GIAO DIỆN CHAT, LỊCH SỬ HIỂN THỊ VÀ NÚT XÓA CHAT
        # =============================================================================
        st.markdown("<p style='font-weight:700; font-size:16px; color:#1E3A8A; margin-top:25px;'>🧠 TRỢ LÝ AI: TỰ ĐỘNG NHẬN DIỆN ĐIỂM ĐO ĐA CHỦNG LOẠI (MÁY TÍNH TỰ TOÁN CHÍNH XÁC)</p>", unsafe_allow_html=True)
        
        if "purchase_chat_history" not in st.session_state:
            st.session_state["purchase_chat_history"] = []
            
        col_chat_header, col_clear_btn = st.columns([4.0, 1.0])
        with col_clear_btn:
            if st.button("🗑️ XÓA LỊCH SỬ CHAT", use_container_width=True, type="secondary", key="clear_consumption_chat_btn"):
                st.session_state["purchase_chat_history"] = []
                st.rerun()
                
        for chat in st.session_state["purchase_chat_history"]:
            with st.chat_message(chat["role"]):
                st.write(chat["content"])
                
                # Hiển thị số liệu định mức trung bình tổng kết lớn do máy tính tính toán chuẩn xác
                if "metric_val" in chat:
                    col_m1, col_m2 = st.columns(2)
                    with col_m1:
                        st.metric(label="🎯 MỐC GỐC PHÂN TÍCH", value=str(chat.get("base_size_lbl", "N/A")))
                    with col_m2:
                        st.metric(label="⚡ ĐỊNH MỨC TRUNG BÌNH TOÀN ĐƠN (SBD)", value=f"{chat['metric_val']:.4f} Yds/Pcs")
                        
                if "df_data" in chat:
                    st.dataframe(chat["df_data"], use_container_width=True, hide_index=True)

        user_msg = st.chat_input("Nhập yêu cầu (Ví dụ: tính định mức theo size chuẩn 32/32 mức 1.73y)...", key="consumption_chat_input_box")
        # =============================================================================
        # ĐOẠN 2: KHỐI XỬ LÝ API BÓC SỐ VÀ THUẬT TOÁN TOÁN HỌC PYTHON CHÍNH XÁC
        # =============================================================================
        if user_msg:
            with st.chat_message("user"):
                st.write(user_msg)
            st.session_state["purchase_chat_history"].append({"role": "user", "content": user_msg})
            
            with st.spinner("🚀 AI đang trích xuất dữ liệu, máy tính đang chạy thuật toán gia quyền chính xác..."):
                if "get_secure_gemini_key" in globals(): gemini_key = get_secure_gemini_key()
                else: gemini_key = st.secrets.get("GEMINI_API_KEY", "").strip()
                
                if not gemini_key:
                    st.error("❌ Không tìm thấy API KEY trong cấu hình hệ thống (Secrets)!")
                    st.stop()
                    
                import json
                import io
                from google import genai
                client_ai_chat = genai.Client(api_key=gemini_key)
                
                sbd_res = st.session_state.get("sbd_parsed_data", {})
                tp_res = st.session_state.get("pur_tp_parsed_data", {})
                raw_list = tp_res.get("data", []) if isinstance(tp_res, dict) else tp_res
                
                # PROMPT: Nghiêm cấm AI làm toán nhân chia, chỉ bốc số thô nạp vào hệ thống
                chat_prompt = f"""You are an advanced data extractor for apparel techpacks.
                The user has given this instruction in Vietnamese: "{user_msg}"
                
                YOUR TASKS:
                1. Based on user request, find the baseline Waist/Size (e.g., 32), baseline Inseam (e.g., 32) and the base consumption (e.g., 1.73).
                2. Identify all relevant POM rows that match the user's keywords (e.g., eo, mông, đùi, gối, nách, ngực...).
                3. Extract the exact numerical spec values for ALL sizes listed in the SBD data. If a POM has multiple sub-rows (like G005), select the correct spec matching that specific size's inseam length.
                
                CRITICAL WRITING RULE:
                Do NOT perform any mathematical calculations for 'final_average_consumption' or 'ĐM Từng Size' yourself. Leave it to the Python engine. You only provide the thô (raw) specifications.

                CURRENT DATA CONTEXT:
                - SBD Data: {json.dumps(sbd_res)}
                - Techpack Spec Data: {json.dumps(raw_list)}
                
                Strictly return a clean raw JSON object with no markdown formatting code blocks:
                {{
                  "explanation": "Giải trình bằng tiếng Việt: Xác nhận các mã vị trí đo đã bóc tách thành công để nạp vào máy tính tính toán.",
                  "base_size_string": "32/32",
                  "base_consumption": 1.73,
                  "extracted_data_rows": [
                    {{
                      "size_combo": "34 X 32",
                      "specs_to_average": [17.25, 21.0, 12.25, 19.75],
                      "text_summary": "Eo: 17.25, Mông: 21.0, Đùi: 12.25, Gối: 19.75",
                      "quantity": 210
                    }}
                  ]
                }}"""
                
                try:
                    res_chat_ai = client_ai_chat.models.generate_content(
                        model='gemini-2.5-flash', 
                        contents=[chat_prompt], 
                        config={"response_mime_type": "application/json"}
                    )
                    
                    clean_res_text = res_chat_ai.text.strip().replace("```json", "").replace("```", "").strip()
                    res_dict = json.loads(clean_res_text)
                    
                    ai_reply = res_dict.get("explanation", "Hệ thống đang đồng bộ dữ liệu toán học.")
                    base_sz_lbl = res_dict.get("base_size_string", "32/32")
                    base_cons = float(res_dict.get("base_consumption", 1.0))
                    extracted_rows = res_dict.get("extracted_data_rows", [])
                    
                    # ⚡ ENGINE TOÁN HỌC PYTHON LÀM VIỆC CHÍNH XÁC TUYỆT ĐỐI
                    base_avg_spec = None
                    for row in extracted_rows:
                        if row.get("size_combo") == base_sz_lbl or row.get("size_combo").replace(" ", "") == base_sz_lbl.replace(" ", ""):
                            specs = row.get("specs_to_average", [])
                            if specs:
                                base_avg_spec = sum([float(x) for x in specs]) / len(specs)
                            break
                    
                    # Dự phòng tìm kiếm mốc lệch tên chuỗi
                    if not base_avg_spec and extracted_rows:
                        base_prefix = base_sz_lbl.split('/')[0].strip() if '/' in base_sz_lbl else base_sz_lbl.strip()
                        for row in extracted_rows:
                            if row.get("size_combo").startswith(base_prefix):
                                specs = row.get("specs_to_average", [])
                                if specs:
                                    base_avg_spec = sum([float(x) for x in specs]) / len(specs)
                                break
                    
                    if not base_avg_spec and extracted_rows:
                        specs = extracted_rows[0].get("specs_to_average", [1.0])
                        base_avg_spec = sum([float(x) for x in specs]) / len(specs)

                    total_fabric_demand = 0.0
                    total_pcs = 0
                    final_table_rows = []
                    
                    for row in extracted_rows:
                        sz_name = row.get("size_combo")
                        specs = row.get("specs_to_average", [])
                        txt_sum = row.get("text_summary", "")
                        qty_int = int(row.get("quantity", 0))
                        
                        if specs and base_avg_spec and base_avg_spec > 0:
                            current_avg_spec = sum([float(x) for x in specs]) / len(specs)
                            size_ratio = current_avg_spec / base_avg_spec
                            size_consumption = base_cons * size_ratio
                        else:
                            size_consumption = base_cons
                            
                        total_fabric_demand += size_consumption * qty_int
                        total_pcs += qty_int
                        
                        final_table_rows.append({
                            "Size Đơn Hàng": sz_name,
                            "Chi Tiết Thông Số Trích Xuất": txt_sum,
                            "ĐM Từng Size (Yds)": round(size_consumption, 4),
                            "Sản Lượng (Pcs)": qty_int
                        })
                        
                    final_average_consumption = total_fabric_demand / total_pcs if total_pcs > 0 else base_cons
                    
                    chat_response_packet = {
                        "role": "assistant", 
                        "content": f"🤖 **Trợ lý AI báo cáo:** {ai_reply}\n\n*(Lưu ý: Toàn bộ bảng số liệu và con số ĐM trung bình dưới đây đã được xử lý qua bộ lõi toán học Python của máy tính để đảm bảo độ chính xác tuyệt đối 100% cho sản xuất).* ",
                        "metric_val": final_average_consumption,
                        "base_size_lbl": base_sz_lbl
                    }
                    
                    if final_table_rows:
                        import pandas as pd
                        chat_response_packet["df_data"] = pd.DataFrame(final_table_rows)
                        
                    st.session_state["purchase_chat_history"].append(chat_response_packet)
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Lỗi xử lý toán học hệ thống: {str(e)}")
                    st.stop()

        st.markdown("<hr style='border:1px solid #E2E8F0; margin-top:20px; margin-bottom:20px;'>", unsafe_allow_html=True)



    # =============================================================================
    # KỊCH BẢN CHỨC NĂNG 2: PHÂN HỆ TÁC NGHIỆP BÀN CẮT ĐA GIÀNG HOÀN CHỈNH
    # =============================================================================
    elif menu_sub.startswith("✂️ CHỨC NĂNG 2"):
        
        # KIỂM TRA ĐIỀU KIỆN 1: Nếu CHƯA bốc tách file SBD thành công
        if not st.session_state.get("purchase_ready"):
            st.markdown("""<div class="card-container"><div class="card-section-header">📋 PHÂN HỆ TÁC NGHIỆP BÀN CẮT ĐA GIÀNG</div>
            <p style="color: #64748B; font-size:13px; margin:0;">Chức năng này không cần thông số rập mẫu. Chỉ cần tải lên File SBD số lượng để máy tính tự động chia tỷ lệ bàn cắt.</p></div>""", unsafe_allow_html=True)
            
            file_sbd_c2 = st.file_uploader("📋 Chọn File SBD Số Lượng Đơn Hàng (Excel/PDF)", type=["xlsx", "xls", "pdf"], key="purchase_sbd_c2_unique")
            
            if file_sbd_c2:
                trigger_btn_c2 = st.button("⚡ SỐ HÓA MA TRẬN SẢN LƯỢNG ĐƠN HÀNG TÁC NGHIỆP", type="primary", use_container_width=True, key="activate_sbd_only_ingest_c2")
                if trigger_btn_c2:
                    with st.spinner("🚀 Hệ thống đang phân tích mảng phân bổ size phẳng từ file SBD..."):
                        if "get_secure_gemini_key" in globals(): 
                            gemini_key = get_secure_gemini_key()
                        else: 
                            gemini_key = st.secrets.get("GEMINI_API_KEY", "").strip()
                        
                        client_ai = genai.Client(api_key=gemini_key)
                        sbd_bytes = file_sbd_c2.getvalue()
                        sbd_content_str = ""
                        sbd_parts_payload = []
                        
                        if file_sbd_c2.name.lower().endswith(('.xlsx', '.xls')):
                            try:
                                excel_data = pd.read_excel(io.BytesIO(sbd_bytes), sheet_name=None)
                                for sheet_name, df_sheet in excel_data.items():
                                    sbd_content_str += f"\n--- SHEET: {sheet_name} ---\n{df_sheet.fillna('').to_csv(index=False)}"
                            except Exception: 
                                pass
                        elif file_sbd_c2.name.lower().endswith('.pdf'):
                            sbd_parts_payload.append(types.Part.from_bytes(data=sbd_bytes, mime_type='application/pdf'))
                            
                        sbd_prompt = "Extract style_id, total_quantity, and flat size mappings. Return raw JSON matching schema: {\"style_id\": \"string\", \"total_quantity\": integer, \"size_breakdown\": {\"Size Name\": integer}}"
                        if sbd_content_str: 
                            sbd_parts_payload.append(types.Part.from_text(text=sbd_content_str))
                        sbd_parts_payload.append(types.Part.from_text(text=sbd_prompt))
                        
                        try:
                            res_sbd = client_ai.models.generate_content(model='gemini-2.5-flash', contents=sbd_parts_payload, config=types.GenerateContentConfig(response_mime_type="application/json"))
                            st.session_state["sbd_parsed_data"] = json.loads(res_sbd.text.strip().replace("```json", "").replace("```", "").strip())
                        except Exception: 
                            pass
                        
                        st.session_state["pur_tp_parsed_data"] = {"dummy_status": "skipped_not_needed"}
                        st.session_state["purchase_ready"] = True
                        st.rerun()
        # KIỂM TRA ĐIỀU KIỆN 2: Nếu ĐÃ số hóa xong file SBD -> Màn hình tác nghiệp sản xuất
        else:
            sbd_data_store = st.session_state.get("sbd_parsed_data", {})
            
            if isinstance(sbd_data_store, dict) and sbd_data_store:
                detected_style_id = sbd_data_store.get("style_id", "UNKNOWN_STYLE")
                detected_total_po = sbd_data_store.get("total_quantity", 0)
                size_breakdown_main = sbd_data_store.get("size_breakdown", {})

                if st.button("🔄 Tải lên File SBD Khác", type="secondary"):
                    st.session_state["purchase_ready"] = False
                    st.session_state["sbd_parsed_data"] = {}
                    st.rerun()

                # KHỐI KHAI BÁO THÔNG SỐ ĐẦU VÀO CỦA MÃ HÀNG HIỆN HÀNH
                st.markdown("#### 📋 KHAI BÁO THÔNG SỐ TÁC NGHIỆP ĐƠN HÀNG VÀ BÀN VẢI MULTI-INSEAM")
                input_col1, input_col2, input_col3 = st.columns(3)
                with input_col1: 
                    style_id_input = st.text_input("🏷️ Tên mã hàng (Style ID):", value=str(detected_style_id).strip().upper())
                with input_col2: 
                    po_qty_input = st.number_input("📦 Số lượng đơn hàng (PO Pcs):", value=int(detected_total_po), step=100)
                with input_col3: 
                    consumption_input = st.number_input("🎯 Định mức tài liệu đề xuất (Yds/Pcs):", value=1.140, step=0.001, format="%.3f")

                input_col4, input_col6 = st.columns(2)
                with input_col4: 
                    max_table_length = st.number_input("📏 Chiều gia tối đa bàn vải (Meters):", value=12.00, step=1.0)
                with input_col6: 
                    cuttable_width_inch = st.number_input("📐 KHỔ CẮT (Khổ vải đi sơ đồ - Inches):", value=56.00, step=0.50, format="%.2f")
                
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("<p style='font-weight:700; font-size:13px; color:#1E3A8A;'>📥 KHU VỰC DÁN DỮ LIỆU CAD (TÊN SƠ ĐỒ & DÀI SƠ ĐỒ COPY TỪ EXCEL)</p>", unsafe_allow_html=True)
                
                cad_paste_zone = st.text_area(
                    "Sau khi xem cấu trúc phối size phía dưới, hãy đi sơ đồ trên máy CAD rồi copy dán kết quả [Tên sơ đồ + Chiều dài mét] vào đây:",
                    placeholder="Ví dụ dán bảng từ Excel CAD:\n5844-c01 1.05\n5844-c02 10", 
                    height=90, 
                    key="cad_bulk_paste_c2"
                )
                
                st.markdown("<p style='font-weight:700; font-size:13px; color:#065F46;'>📊 MA TRẬN SẢN LƯỢNG ĐƠN HÀNG (SIZE BREAKDOWN TỪ SBD)</p>", unsafe_allow_html=True)
                if size_breakdown_main:
                    df_size = pd.DataFrame([size_breakdown_main])
                    st.dataframe(df_size, use_container_width=True, hide_index=True)
                
                # --- TRUNG TÂM TRA CỨU DATABASE SUPABASE ---
                st.markdown("<p style='font-weight:700; font-size:13px; color:#1E3A8A; margin-top:15px;'>🔍 TRUNG TÂM TRA CỨU DATABASE SUPABASE</p>", unsafe_allow_html=True)
                db_search_query = st.text_input("Tìm kiếm mã hàng đã tác nghiệp trên hệ thống Supabase:", placeholder="Nhập Style Name để gọi lại thông số cũ...", key="subapat_db_search")
                if db_search_query.strip():
                    try:
                        search_res = st.session_state.supabase.table("tac_nghiep_ban_cat").select("*").eq("style_name", db_search_query.strip().upper()).execute()
                        if search_res.data:
                            st.success(f"📌 Tìm thấy dữ liệu lịch sử của mã hàng {db_search_query.strip().upper()} trên Supabase!")
                            matched_row = search_res.data
                            st.info(f"Sản lượng cũ: {matched_row.get('po_quantity')} Pcs | Định mức cũ: {matched_row.get('consumption_value')} Yds")
                    except Exception:
                        pass

                # Định nghĩa cấu trúc mảng kích cỡ và nhóm Inseam
                active_sizes = [str(k) for k, v in size_breakdown_main.items() if int(v) > 0]
                if not active_sizes:
                    active_sizes = ["S", "M", "L", "XL", "2XL", "3XL"]
                detected_inseam = sbd_data_store.get("inseam_group", "None")
                st.markdown(f"**📌 Nhóm Inseam hiện hành:** `{detected_inseam}`")
                
                # Bộ đôi hai nút bấm kích hoạt quy trình sản xuất độc lập
                btn_col1, btn_col2 = st.columns(2)
                with btn_col1:
                    trigger_auto_cutting = st.button("⚡ 1. KÍCH HOẠT TÍNH TÁC NGHIỆP SƠ ĐỒ (HÌNH THÁP)", type="primary", use_container_width=True)
                with btn_col2:
                    trigger_consumption = st.button("📊 2. KÍCH HOẠT TÍNH ĐỊNH MỨC (KHI ĐÃ CÓ CAD)", type="secondary", use_container_width=True)

                if "auto_cutting_results" not in st.session_state:
                    st.session_state["auto_cutting_results"] = None
                if "consumption_activated" not in st.session_state:
                    st.session_state["consumption_activated"] = False

                                # =============================================================================
                                # =============================================================================
                                # =============================================================================
                                # =============================================================================
                # ĐOẠN 1: THUẬT TOÁN HÌNH THÁP NGƯỢC TỰ ĐỘNG THU NGẮN SƠ ĐỒ ĐỂ TĂNG SỐ LỚP VẢI
                # =============================================================================
                st.markdown("<br>", unsafe_allow_html=True)
                
                active_sizes = [str(k) for k, v in size_breakdown_main.items() if int(v) > 0]
                if not active_sizes:
                    active_sizes = ["S", "M", "L", "XL", "2XL", "3XL"]
                
                detected_inseam = sbd_data_store.get("inseam_group", "None")
                st.markdown(f"**📌 Nhóm Inseam hiện hành:** `{detected_inseam}`")
                
                btn_col1, btn_col2 = st.columns(2)
                with btn_col1:
                    trigger_auto_cutting = st.button("⚡ 1. KÍCH HOẠT TÍNH TÁC NGHIỆP SƠ ĐỒ (HÌNH THÁP)", type="primary", use_container_width=True, key="c2_btn_auto_cut_unique")
                with btn_col2:
                    trigger_consumption = st.button("📊 2. KÍCH HOẠT TÍNH ĐỊNH MỨC (KHI ĐÃ CÓ CAD)", type="secondary", use_container_width=True, key="c2_btn_consumption_unique")

                if "auto_cutting_results" not in st.session_state:
                    st.session_state["auto_cutting_results"] = None
                if "consumption_activated" not in st.session_state:
                    st.session_state["consumption_activated"] = False

                if trigger_auto_cutting:
                    st.session_state["consumption_activated"] = False
                    with st.spinner("🔮 Hệ thống đang tối ưu hóa số lớp và chiều dài sơ đồ..."):
                        import math
                        cons_meters = consumption_input / 1.09361
                        max_pcs_per_marker = math.floor(max_table_length / (cons_meters if cons_meters > 0 else 1.0))
                        if max_pcs_per_marker <= 0: max_pcs_per_marker = 6
                        
                        balance_tracker = {sz: int(size_breakdown_main.get(sz, 0)) for sz in active_sizes}
                        calculated_steps = []
                        step_idx = 1
                        max_loops = 25
                        
                        while sum(balance_tracker.values()) > 0 and step_idx <= max_loops:
                            marker_id = f"c{step_idx:02d}"
                            
                            # Cấu hình mốc lớp tiêu chuẩn cho các tầng hình tháp
                            if step_idx == 1: target_layers = 150
                            elif step_idx == 2: target_layers = 120
                            elif step_idx == 3: target_layers = 90
                            else: target_layers = 60
                            
                            sorted_sizes = sorted(balance_tracker.items(), key=lambda x: x, reverse=True)
                            current_ratios = {sz: 0 for sz in active_sizes}
                            assigned_pcs = 0
                            
                            # 🎯 CẢI TIẾN CỐT LÕI: Kiểm tra sản lượng còn lại lớn nhất
                            max_remaining_bal = max(balance_tracker.values()) if balance_tracker.values() else 0
                            
                            # NGUYÊN TẮC MAY: Nếu lượng hàng tồn kho còn quá mỏng, ép thu ngắn chiều dài sơ đồ lại
                            effective_max_pcs = max_pcs_per_marker
                            if max_remaining_bal < target_layers and max_remaining_bal > 0:
                                # Nếu sản lượng quá ít, tự động ép sơ đồ ngắn chỉ đi 1 đến 2 sản phẩm
                                effective_max_pcs = min(2, max_pcs_per_marker)
                                target_layers = max_remaining_bal # Ép số lớp đạt đỉnh sản lượng để triệt tiêu
                            
                            for sz, bal in sorted_sizes:
                                if bal <= 0 or assigned_pcs >= effective_max_pcs: 
                                    continue
                                
                                needed_ratio = math.floor(bal / target_layers)
                                if needed_ratio > 4: needed_ratio = 4
                                if needed_ratio == 0 and bal > (target_layers / 2): 
                                    needed_ratio = 1
                                    
                                if assigned_pcs + needed_ratio > effective_max_pcs:
                                    needed_ratio = effective_max_pcs - assigned_pcs
                                    
                                current_ratios[sz] = needed_ratio
                                assigned_pcs += needed_ratio
                            
                            if assigned_pcs == 0:
                                # Gom nốt lượng hàng lẻ ở đuôi đơn hàng vào sơ đồ cực ngắn
                                effective_max_pcs = min(2, max_pcs_per_marker)
                                for sz, bal in sorted_sizes:
                                    if bal > 0 and assigned_pcs < effective_max_pcs:
                                        current_ratios[sz] = 1
                                        assigned_pcs += 1
                                        
                            # Tính toán số lớp thực tế dựa trên tỷ lệ mới phối
                            layer_candidates = []
                            for sz in active_sizes:
                                rat = current_ratios[sz]
                                if rat > 0:
                                    layer_candidates.append(math.ceil(balance_tracker[sz] / rat))
                            
                            computed_layers = min(layer_candidates) if layer_candidates else target_layers
                            if computed_layers <= 0: computed_layers = 1
                            
                            # Cân đối bàn cắt thương mại
                            if computed_layers > 150:
                                num_tables = math.ceil(computed_layers / 120)
                                computed_layers = math.ceil(computed_layers / num_tables)
                            else:
                                num_tables = 1
                                
                            calculated_steps.append({
                                "Sơ đồ / Trạng thái": marker_id,
                                "Số lớp": computed_layers, "Số bàn": num_tables,
                                "Dài sơ đồ": 0.0, "Số sp/SĐ": assigned_pcs, "Ratios": current_ratios
                            })
                            
                            for sz in active_sizes:
                                total_cut = current_ratios[sz] * computed_layers * num_tables
                                balance_tracker[sz] = max(0, balance_tracker[sz] - total_cut)
                                
                            calculated_steps.append({
                                "Sơ đồ / Trạng thái": "Balance",
                                "Số lớp": "", "Số bàn": "", "Dài sơ đồ": "", "Số sp/SĐ": "",
                                "Ratios": {sz: balance_tracker[sz] for sz in active_sizes}
                            })
                            step_idx += 1
                            
                        st.session_state["auto_cutting_results"] = calculated_steps

                if trigger_consumption:
                    st.session_state["consumption_activated"] = True
                    st.rerun()




                               # BƯỚC 3: LIÊN KẾT ĐỐI CHIẾU DỮ LIỆU Ô CAD, ĐẨY SUPABASE & KẾT XUẤT EXCEL
                if st.session_state.get("auto_cutting_results") is not None:
                    import re
                    import io
                    
                    cad_lengths_map = {}
                    if cad_paste_zone.strip() and st.session_state["consumption_activated"]:
                        cad_lines = cad_paste_zone.strip().split("\n")
                        for line in cad_lines:
                            if not line.strip(): continue
                            match = re.search(r'(c\d{2})[\s\t]+([0-9]*\.?[0-9]+)', line.lower().strip())
                            if match:
                                suffix_key = match.group(1)
                                try:
                                    cad_lengths_map[suffix_key] = float(match.group(2))
                                except ValueError: pass

                    final_rows_display = []
                    total_fabric_m = 0.0
                    total_cut_pcs_sum = 0
                    
                    for item in st.session_state["auto_cutting_results"]:
                        display_row = {"SIZE": item["Sơ đồ / Trạng thái"]}
                        for sz in active_sizes:
                            display_row[sz] = item["Ratios"].get(sz, 0)
                            
                        if item["Sơ đồ / Trạng thái"] != "Balance":
                            layers = item["Số lớp"]
                            tables = item["Số bàn"]
                            sp_sd = item["Số sp/SĐ"]
                            current_marker_id = item["Sơ đồ / Trạng thái"].lower().strip()
                            
                            m_len = cad_lengths_map.get(current_marker_id, 0.0) if st.session_state["consumption_activated"] else 0.0
                            vail_can_m = m_len * layers * tables
                            total_fabric_m += vail_can_m
                            
                            total_ratios_sum = sum(item["Ratios"].values())
                            pcs_cut = total_ratios_sum * layers * tables
                            total_cut_pcs_sum += pcs_cut
                            
                            dm_sd = (vail_can_m * 1.09361) / pcs_cut if pcs_cut > 0 else 0.0
                            
                            display_row["Số lớp"] = layers
                            display_row["Số bàn"] = tables
                            display_row["Dài sơ đồ"] = m_len
                            display_row["Số sp/SĐ"] = sp_sd
                            display_row["Đ.Mức SĐ"] = round(dm_sd, 3)
                            display_row["Vải cần (M)"] = round(vail_can_m, 1)
                        else:
                            display_row["Số lớp"] = ""
                            display_row["Số bàn"] = ""
                            display_row["Dài sơ đồ"] = ""
                            display_row["Số sp/SĐ"] = ""
                            display_row["Đ.Mức SĐ"] = ""
                            display_row["Vải cần (M)"] = ""
                        final_rows_display.append(display_row)
                        
                    df_final_report = pd.DataFrame(final_rows_display)
                    total_fabric_yds_final = total_fabric_m * 1.09361
                    final_avg_yield = total_fabric_yds_final / (total_cut_pcs_sum if total_cut_pcs_sum > 0 else 1)
                    
                    if st.button("💾 ĐẨY DỮ LIỆU TÁC NGHIỆP LÊN DATABASE SUPABASE", type="secondary", use_container_width=True):
                        try:
                            payload_db = {
                                "style_name": str(style_id_input).strip().upper(), "po_quantity": int(po_qty_input),
                                "planned_cut_pcs": int(total_cut_pcs_sum), "consumption_value": str(round(final_avg_yield, 3)),
                                "total_material_value": str(round(total_fabric_yds_final, 2)), "cuttable_width_inch": float(cuttable_width_inch)
                            }
                            if "get_supabase_client" in globals():
                                sb_client = get_supabase_client()
                                sb_client.table("tac_nghiep_ban_cat").insert(payload_db).execute()
                                st.success(f"🎉 Đã đồng bộ dữ liệu mã hàng {style_id_input} lên hệ thống Supabase thành công!")
                            else:
                                st.error("❌ Hệ thống chưa khởi tạo kết nối database. Vui lòng bấm F5 tải lại trang.")
                        except Exception as db_err:
                            st.error(f"Lỗi khi đẩy dữ liệu lên Supabase: {str(db_err)}")
                    try:
                        buffer = io.BytesIO()
                        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                            header_data = {
                                "THÔNG TIN ĐƠN HÀNG TÁC NGHIỆP BÀN CẮT": [
                                    f"Mã hàng (Style Name): {style_id_input}", f"Số lượng đơn hàng (PO Qty): {po_qty_input} Pcs",
                                    f"Sản lượng kế hoạch cắt (Planned Cut): {total_cut_pcs_sum} Pcs", f"Định mức tài liệu đề xuất: {consumption_input:.3f} Yds/Pcs",
                                    f"Định mức tác nghiệp thực tế: {final_avg_yield:.3f} Yds/Pcs", f"Khổ vải đi sơ đồ (Inches): {cuttable_width_inch}\"",
                                    f"Nhóm phân hệ Inseam: {detected_inseam}"
                                ]
                            }
                            pd.DataFrame(header_data).to_excel(writer, sheet_name="BaoCao_TacNghiep", index=False, startrow=0)
                            df_final_report.to_excel(writer, sheet_name="BaoCao_TacNghiep", index=False, startrow=10)
                            
                            worksheet = writer.sheets["BaoCao_TacNghiep"]
                            from openpyxl.styles import PatternFill, Font
                            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
                            red_font = Font(color="991B1B", bold=True)
                            
                            for r_idx in range(12, worksheet.max_row + 1):
                                if worksheet.cell(row=r_idx, column=1).value == "Balance":
                                    for c_idx in range(1, worksheet.max_column + 1):
                                        cell = worksheet.cell(row=r_idx, column=c_idx)
                                        cell.fill = yellow_fill
                                        cell.font = red_font
                        
                        st.download_button(
                            label="📥 XUẤT FILE EXCEL TÁC NGHIỆP CHUẨN THƯƠNG MẠI", data=buffer.getvalue(),
                            file_name=f"BÁO_CÁO_TÁC_NGHIỆP_BÀN_CẮT_{style_id_input}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True
                        )
                    except Exception: pass

                    parsed_size_columns = []
                    for col_name in active_sizes:
                        col_str = str(col_name).strip()
                        if any(char in col_str.lower() for char in ["x", "-", "/"]):
                            parts = re.split(r'[\sXx\-\/]+', col_str)
                            if len(parts) >= 2:
                                parsed_size_columns.append({
                                    "original": col_name, 
                                    "size_num": str(parts[0]).strip(), 
                                    "giang_num": str(parts[1]).strip()
                                })
                            else:
                                parsed_size_columns.append({"original": col_name, "size_num": col_str, "giang_num": str(detected_inseam)})
                        else:
                            parsed_size_columns.append({"original": col_name, "size_num": col_str, "giang_num": str(detected_inseam)})

                    try:
                        parsed_size_columns.sort(key=lambda x: (int(re.sub(r'\D', '', x['giang_num'])), int(re.sub(r'\D', '', x['size_num']))))
                    except Exception:
                        parsed_size_columns.sort(key=lambda x: (x['giang_num'], x['size_num']))

                    ordered_size_keys = [item["original"] for item in parsed_size_columns]
                    other_tech_keys = ["Số lớp", "Số bàn", "Dài sơ đồ", "Số sp/SĐ", "Đ.Mức SĐ", "Vải cần (M)"]
                    df_final_report = df_final_report[["SIZE"] + ordered_size_keys + other_tech_keys]

                    multi_cols_tuples = [("", "SIZE")]
                    for item in parsed_size_columns:
                        multi_cols_tuples.append((f"GIÀNG: {item['giang_num']}", item['size_num']))
                    for col_name in other_tech_keys:
                        multi_cols_tuples.append(("THÔNG SỐ TÁC NGHIỆP", col_name))
                    
                    df_final_report.columns = pd.MultiIndex.from_tuples(multi_cols_tuples)
                    def style_full_balance_rows(row):
                        if row.iloc[0] == "Balance":
                            return ['background-color: #FEF08A; color: #991B1B; font-weight: 700; border: 1px solid #FDE047;'] * len(row)
                        return [''] * len(row)
                    
                    styled_df_report = df_final_report.style.apply(style_full_balance_rows, axis=1)

                    st.markdown("<p style='font-weight:700; font-size:14px; color:#1E3A8A; margin-top:15px;'>📊 BẢNG THEO DÕI TÁC NGHIỆP & CÂN ĐỐI ĐƠN HÀNG MULTI-INSEAM</p>", unsafe_allow_html=True)
                    st.dataframe(styled_df_report, use_container_width=True, hide_index=True)
                    
                    st.markdown("---")
                    m_col1, m_col2, m_col3 = st.columns(3)
                    with m_col1:
                        st.metric("Tổng vải tiêu thụ tự động", f"{total_fabric_m:,.1f} Mét")
                    with m_col2:
                        st.metric("Định mức trung bình (Đ.Mức TB)", f"{final_avg_yield:.3f} Yds/Pcs" if st.session_state["consumption_activated"] else "0.000 Yds/Pcs")
                    with m_col3:
                        variance = final_avg_yield - consumption_input if total_fabric_m > 0 and st.session_state["consumption_activated"] else 0.0
                        st.metric("Chênh lệch so với tài liệu", f"{variance:+.3f}" if st.session_state["consumption_activated"] else "0.000", delta_color="inverse" if variance > 0 else "normal")
                else:
                    st.info("💡 Quy trình: Bấm nút 1 để tính tác nghiệp sơ đồ -> Điền độ dài CAD -> Bấm nút 2 để kích hoạt nhảy số định mức.")
